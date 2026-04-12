"""
EVO MDD Verification Tool — Main Launcher
==========================================
Philips light theme. Two visible tools:
  Step 1 — ISO Symbol Library (upload PDF → silent backend extraction)
  Step 2 — Start Verification (EVO MDD Verification Tool)

V&V Plan upload is integrated directly inside the Verification window.
The GTIN/Symbol Mapper has been removed — all requirements come from the V&V Plan.
"""

import logging
import os
import sys
import threading
import tkinter as tk
from tkinter import ttk, messagebox

_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

# ── Logging ────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s %(name)s: %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(_HERE, 'error.log'), encoding='utf-8'),
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)
for _noisy in ('PIL', 'PIL.PngImagePlugin', 'PIL.TiffImagePlugin', 'PIL.JpegImagePlugin'):
    logging.getLogger(_noisy).setLevel(logging.WARNING)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED STATE BUS
# ══════════════════════════════════════════════════════════════════════════════

APP_STATE: dict = {
    "icon_library_path": "",
    "required_symbols":  [],
    "verify_list":       [],
    "vv_plan_path":      "",
    "vv_plan_data":      None,
    "callbacks": {
        "icon_library_ready": [],
        "symbols_ready":      [],
        "vv_plan_ready":      [],
    },
}


def register_callback(event: str, fn) -> None:
    APP_STATE["callbacks"].setdefault(event, []).append(fn)


def fire_event(event: str, *args) -> None:
    for fn in APP_STATE["callbacks"].get(event, []):
        try:
            fn(*args)
        except Exception:
            logger.exception("Callback error for event '%s'", event)


# ══════════════════════════════════════════════════════════════════════════════
# PHILIPS LIGHT THEME — Design tokens
# ══════════════════════════════════════════════════════════════════════════════

COLORS: dict = {
    # Backgrounds
    "bg":           "#F5F7FA",
    "bg_card":      "#FFFFFF",
    "bg_input":     "#FFFFFF",
    # Philips blues
    "accent":       "#0B54A4",      # Philips primary blue
    "accent_light": "#00A0DC",      # Philips light blue
    "accent2":      "#16A34A",      # success green
    "accent3":      "#D97706",      # amber / warning
    # Text
    "text":         "#1A1A2E",
    "text_muted":   "#6B7280",
    # Borders & states
    "border":       "#D1D5DB",
    "header_bg":    "#0B54A4",
    "danger":       "#DC2626",
    "success":      "#16A34A",
    "warn":         "#D97706",
    # Log area — light, readable
    "log_bg":       "#F8FAFC",
    "log_text":     "#1E3A5F",
    # Sidebar
    "sidebar":      "#0B2B5C",
    "sidebar_hover":"#0B54A4",
}

FONT_TITLE   = ("Segoe UI", 16, "bold")
FONT_HEADING = ("Segoe UI", 11, "bold")
FONT_BODY    = ("Segoe UI",  9)
FONT_MONO    = ("Consolas",  9)
FONT_SMALL   = ("Segoe UI",  8)


def _lighten(hex_color: str) -> str:
    hex_color = hex_color.lstrip("#")
    r, g, b = [int(hex_color[i:i+2], 16) for i in (0, 2, 4)]
    return f"#{min(255,r+30):02x}{min(255,g+30):02x}{min(255,b+30):02x}"


def _darken(hex_color: str) -> str:
    hex_color = hex_color.lstrip("#")
    r, g, b = [int(hex_color[i:i+2], 16) for i in (0, 2, 4)]
    return f"#{max(0,r-20):02x}{max(0,g-20):02x}{max(0,b-20):02x}"


# ══════════════════════════════════════════════════════════════════════════════
# SHARED WIDGET FACTORIES
# ══════════════════════════════════════════════════════════════════════════════

def styled_button(parent, text="", command=None, color=None, **kwargs) -> tk.Button:
    bg = color or COLORS["accent"]
    return tk.Button(
        parent, text=text, command=command,
        bg=bg, fg="white",
        activebackground=_darken(bg), activeforeground="white",
        relief="flat", cursor="hand2",
        font=FONT_BODY, padx=12, pady=6, bd=0, **kwargs,
    )


def styled_label(parent, text="", muted=False, **kwargs) -> tk.Label:
    fg = COLORS["text_muted"] if muted else COLORS["text"]
    return tk.Label(parent, text=text, bg=COLORS["bg_card"], fg=fg,
                    font=FONT_BODY, **kwargs)


def styled_entry(parent, textvariable=None, readonly=False, **kwargs) -> tk.Entry:
    state = "readonly" if readonly else "normal"
    return tk.Entry(
        parent, textvariable=textvariable, state=state,
        bg=COLORS["bg_input"], fg=COLORS["text"],
        readonlybackground=COLORS["bg_input"],
        insertbackground=COLORS["text"],
        relief="flat", font=FONT_MONO,
        highlightthickness=1, highlightbackground=COLORS["border"], **kwargs,
    )


# ══════════════════════════════════════════════════════════════════════════════
# PROGRESS POPUP  (theme-aware)
# ══════════════════════════════════════════════════════════════════════════════

class ProgressPopup(tk.Toplevel):
    _PENDING_FG = COLORS["text_muted"]
    _ACTIVE_FG  = COLORS["accent"]
    _DONE_FG    = COLORS["accent2"]
    _ERROR_FG   = COLORS["danger"]

    def __init__(self, parent, title: str = "Processing", steps: list = None):
        super().__init__(parent)
        self.title(title)
        self.configure(bg=COLORS["bg"])
        self.resizable(False, False)
        self.grab_set()

        self._steps      = steps or []
        self._step_vars  = []
        self._detail_var = tk.StringVar(value="Starting…")
        self._status_var = tk.StringVar(value="")

        self._build_ui(title)
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        w, h = 480, 320
        try:
            px = parent.winfo_rootx() + parent.winfo_width()  // 2
            py = parent.winfo_rooty() + parent.winfo_height() // 2
        except Exception:
            px, py = 400, 300
        self.geometry(f"{w}x{h}+{px - w//2}+{py - h//2}")

    def _build_ui(self, title: str):
        hdr = tk.Frame(self, bg=COLORS["header_bg"], padx=16, pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=title, bg=COLORS["header_bg"], fg="white",
                 font=FONT_HEADING).pack(anchor="w")

        steps_frame = tk.Frame(self, bg=COLORS["bg_card"], padx=16, pady=12)
        steps_frame.pack(fill="x", padx=8, pady=(8, 0))

        for i, step_text in enumerate(self._steps):
            var = tk.StringVar(value=f"○  {step_text}")
            self._step_vars.append(var)
            lbl = tk.Label(steps_frame, textvariable=var,
                           bg=COLORS["bg_card"], fg=self._PENDING_FG,
                           font=FONT_BODY, anchor="w")
            lbl.pack(fill="x", pady=1)
            self._step_vars[i] = (var, lbl, step_text)

        detail_frame = tk.Frame(self, bg=COLORS["bg"], padx=16, pady=6)
        detail_frame.pack(fill="x", padx=8)
        tk.Label(detail_frame, textvariable=self._detail_var,
                 bg=COLORS["bg"], fg=COLORS["text_muted"],
                 font=FONT_SMALL, anchor="w", wraplength=440).pack(fill="x")

        pb_frame = tk.Frame(self, bg=COLORS["bg"], padx=16, pady=4)
        pb_frame.pack(fill="x", padx=8)
        style = ttk.Style(self)
        style.configure("Popup.Horizontal.TProgressbar",
                        troughcolor=COLORS["border"],
                        background=COLORS["accent"], thickness=6)
        self._pb = ttk.Progressbar(pb_frame, orient="horizontal",
                                   mode="indeterminate",
                                   style="Popup.Horizontal.TProgressbar")
        self._pb.pack(fill="x")
        self._pb.start(20)

        bottom = tk.Frame(self, bg=COLORS["bg"], padx=16, pady=8)
        bottom.pack(fill="x", padx=8, side="bottom")
        tk.Label(bottom, textvariable=self._status_var,
                 bg=COLORS["bg"], fg=COLORS["text_muted"],
                 font=FONT_SMALL, anchor="w").pack(side="left", fill="x", expand=True)
        self._close_btn = tk.Button(
            bottom, text="Close", command=self.destroy,
            bg=COLORS["border"], fg=COLORS["text"],
            activebackground=_lighten(COLORS["border"]),
            activeforeground=COLORS["text"],
            relief="flat", cursor="hand2",
            font=FONT_SMALL, padx=10, pady=4, bd=0, state="disabled",
        )
        self._close_btn.pack(side="right")

    def set_step(self, index: int, detail: str = ""):
        for i, (var, lbl, text) in enumerate(self._step_vars):
            if i < index:
                var.set(f"✔  {text}"); lbl.configure(fg=self._DONE_FG)
            elif i == index:
                var.set(f"▶  {text}"); lbl.configure(fg=self._ACTIVE_FG)
            else:
                var.set(f"○  {text}"); lbl.configure(fg=self._PENDING_FG)
        if detail:
            self._detail_var.set(detail)
        self.update_idletasks()

    def set_detail(self, detail: str):
        self._detail_var.set(detail)
        self.update_idletasks()

    def finish(self, msg: str = "Complete."):
        for var, lbl, text in self._step_vars:
            var.set(f"✔  {text}"); lbl.configure(fg=self._DONE_FG)
        self._detail_var.set(msg)
        self._status_var.set("Finished")
        self._pb.stop()
        self._pb.configure(mode="determinate")
        ttk.Style(self).configure("Popup.Horizontal.TProgressbar",
                                  background=COLORS["accent2"])
        self._pb["value"] = 100
        self._close_btn.configure(state="normal", bg=COLORS["accent2"],
                                  fg="white")
        self.update_idletasks()

    def error(self, msg: str = "An error occurred."):
        self._detail_var.set(f"Error: {msg}")
        self._status_var.set("Failed")
        self._pb.stop()
        self._pb.configure(mode="determinate")
        ttk.Style(self).configure("Popup.Horizontal.TProgressbar",
                                  background=COLORS["danger"])
        self._pb["value"] = 100
        self._close_btn.configure(state="normal", bg=COLORS["danger"], fg="white")
        for var, lbl, text in self._step_vars:
            if var.get().startswith("▶"):
                lbl.configure(fg=self._ERROR_FG)
        self.update_idletasks()


# ══════════════════════════════════════════════════════════════════════════════
# ISO SYMBOL LIBRARY — Simple upload screen (Step 2)
# Replaces the full ExtractorWindow — just Browse → Progress → Done
# ══════════════════════════════════════════════════════════════════════════════

_ICON_LIBRARY_DIR = os.path.join(_HERE, "Icon Library")


def _library_is_built() -> bool:
    """Return True if the Icon Library already contains extracted symbols."""
    json_path = os.path.join(_ICON_LIBRARY_DIR, "iso_symbols.json")
    return os.path.isfile(json_path) and len(
        [f for f in os.listdir(_ICON_LIBRARY_DIR)
         if f.lower().endswith(".png")]
    ) > 0 if os.path.isdir(_ICON_LIBRARY_DIR) else False


class ISOLibraryUploadWindow(tk.Toplevel):
    """
    Simple one-screen tool:
      Upload ISO 15223 PDF → progress bar → extraction done silently.
    No row-coordinate details, no log noise, no gallery.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("ISO Symbol Library — EVO MDD Verification Tool")
        self.configure(bg=COLORS["bg"])
        self.resizable(False, False)
        self._pdf_path = tk.StringVar()
        self._status   = tk.StringVar(value="")
        self._build_ui()
        self._center()
        # If already built, show that status immediately
        if _library_is_built():
            n = len([f for f in os.listdir(_ICON_LIBRARY_DIR)
                     if f.lower().endswith(".png")])
            self._set_done(f"Symbol library is ready — {n} symbols available.")

    def _center(self):
        self.update_idletasks()
        w, h = 560, 380
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self, bg=COLORS["header_bg"], padx=20, pady=16)
        hdr.pack(fill="x")
        tk.Label(hdr, text="ISO Symbol Library",
                 bg=COLORS["header_bg"], fg="white",
                 font=FONT_TITLE).pack(anchor="w")
        tk.Label(hdr, text="Upload the ISO 15223 standard PDF — symbols are extracted automatically",
                 bg=COLORS["header_bg"], fg="#BFDBFE",
                 font=FONT_SMALL).pack(anchor="w", pady=(4, 0))

        body = tk.Frame(self, bg=COLORS["bg"], padx=28, pady=24)
        body.pack(fill="both", expand=True)

        # Instruction
        tk.Label(body,
                 text="The symbol reference library is built once from the ISO 15223 standard.\n"
                      "Upload the PDF below and the extraction runs automatically in the background.",
                 bg=COLORS["bg"], fg=COLORS["text"],
                 font=FONT_BODY, justify="left", wraplength=500).pack(anchor="w", pady=(0, 20))

        # File picker row
        pick_frame = tk.Frame(body, bg=COLORS["bg"])
        pick_frame.pack(fill="x", pady=(0, 6))
        tk.Label(pick_frame, text="ISO 15223 Standard PDF",
                 bg=COLORS["bg"], fg=COLORS["text_muted"],
                 font=FONT_SMALL).pack(anchor="w", pady=(0, 4))

        row = tk.Frame(pick_frame, bg=COLORS["bg"])
        row.pack(fill="x")
        entry = tk.Entry(row, textvariable=self._pdf_path, state="readonly",
                         bg=COLORS["bg_input"], fg=COLORS["text"],
                         readonlybackground=COLORS["bg_input"],
                         relief="flat", font=FONT_MONO,
                         highlightthickness=1,
                         highlightbackground=COLORS["border"])
        entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Button(row, text="Browse…", command=self._browse,
                  bg=COLORS["bg_card"], fg=COLORS["text"],
                  activebackground=COLORS["border"],
                  relief="flat", cursor="hand2",
                  font=FONT_SMALL, padx=10, pady=5, bd=0,
                  highlightthickness=1,
                  highlightbackground=COLORS["border"]).pack(side="left")

        # Progress bar
        self._pb_var = tk.DoubleVar(value=0)
        self._pb = ttk.Progressbar(body, variable=self._pb_var,
                                   maximum=100, mode="indeterminate")
        self._pb.pack(fill="x", pady=(14, 4))

        # Status label
        self._status_lbl = tk.Label(body, textvariable=self._status,
                                     bg=COLORS["bg"], fg=COLORS["text_muted"],
                                     font=FONT_SMALL, anchor="w")
        self._status_lbl.pack(anchor="w")

        # Extract button
        btn_row = tk.Frame(body, bg=COLORS["bg"])
        btn_row.pack(fill="x", pady=(20, 0))
        self._run_btn = tk.Button(
            btn_row, text="Build Symbol Library",
            command=self._run,
            bg=COLORS["accent"], fg="white",
            activebackground=_darken(COLORS["accent"]),
            activeforeground="white",
            relief="flat", cursor="hand2",
            font=("Segoe UI", 10, "bold"),
            padx=20, pady=10, bd=0,
        )
        self._run_btn.pack(side="left")

        tk.Label(btn_row,
                 text="This only needs to be done once.\nThe library is reused across all sessions.",
                 bg=COLORS["bg"], fg=COLORS["text_muted"],
                 font=FONT_SMALL, justify="left").pack(side="left", padx=16)

    def _refocus(self):
        self.lift(); self.focus_force()
        self.attributes("-topmost", True)
        self.after(150, lambda: self.attributes("-topmost", False))

    def _browse(self):
        from tkinter import filedialog
        p = filedialog.askopenfilename(
            parent=self,
            title="Select ISO 15223 Standard PDF",
            filetypes=[("PDF files", "*.pdf *.PDF")])
        if p:
            self._pdf_path.set(p)
            self._status.set("PDF selected. Click 'Build Symbol Library' to start.")
            self._status_lbl.configure(fg=COLORS["text_muted"])
        self._refocus()

    def _run(self):
        if not self._pdf_path.get():
            messagebox.showwarning("No PDF Selected",
                                   "Please browse for the ISO 15223 standard PDF first.",
                                   parent=self)
            return
        self._run_btn.configure(state="disabled", text="Building…")
        self._pb.configure(mode="indeterminate")
        self._pb.start(15)
        self._status.set("Extracting symbols from ISO 15223 PDF…")
        self._status_lbl.configure(fg=COLORS["accent"])
        threading.Thread(target=self._extract_thread, daemon=True).start()

    def _extract_thread(self):
        try:
            from iso_symbol_extractor import extract_symbols_from_pdf, save_json_csv
            os.makedirs(_ICON_LIBRARY_DIR, exist_ok=True)

            def _silent_log(msg, level="INFO"):
                pass   # suppress all log noise

            data = extract_symbols_from_pdf(
                self._pdf_path.get(),
                _ICON_LIBRARY_DIR,
                _silent_log,
                progress_fn=None,
            )
            save_json_csv(data, _ICON_LIBRARY_DIR, _silent_log)

            # Update APP_STATE and fire event
            APP_STATE["icon_library_path"] = _ICON_LIBRARY_DIR
            fire_event("icon_library_ready", _ICON_LIBRARY_DIR)

            self.after(0, self._set_done,
                       f"Done — {len(data)} symbols extracted and saved to the library.")

        except Exception as e:
            import traceback
            logger.exception("ISO extraction failed")
            self.after(0, self._set_error, str(e))

    def _set_done(self, msg: str):
        self._pb.stop()
        self._pb.configure(mode="determinate")
        self._pb_var.set(100)
        self._status.set(f"✔  {msg}")
        self._status_lbl.configure(fg=COLORS["success"])
        self._run_btn.configure(
            state="normal", text="Rebuild Library",
            bg=COLORS["accent2"],
        )

    def _set_error(self, msg: str):
        self._pb.stop()
        self._pb_var.set(0)
        self._status.set(f"✗  Error: {msg}")
        self._status_lbl.configure(fg=COLORS["danger"])
        self._run_btn.configure(state="normal", text="Build Symbol Library",
                                bg=COLORS["accent"])


# ══════════════════════════════════════════════════════════════════════════════
# V&V PLAN UPLOAD WINDOW — one-time setup (shared by both verifications)
# ══════════════════════════════════════════════════════════════════════════════

class VVPlanWindow(tk.Toplevel):
    """Load V&V Plan .docx, display extracted requirements, export to Excel.

    Parsed data stored in APP_STATE["vv_plan_data"] and shared automatically
    with both Label Verification and IFU Verification windows.
    """

    _COL_IDS  = ("prd_id", "type", "vv_method", "requirement_text")
    _COL_HDRS = ("PRD ID", "Type", "V&V Method", "Requirement Text")
    _COL_W    = (90, 60, 100, 500)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("V&V Plan — EVO MDD Verification Tool")
        self.configure(bg=COLORS["bg_card"])
        self.resizable(True, True)
        self._path_var    = tk.StringVar()
        self._status_var  = tk.StringVar(value="")
        self._plan_data   = None   # holds VVPlanData after successful parse
        self._build_ui()
        self._center()
        # Pre-fill if already loaded
        existing_data = APP_STATE.get("vv_plan_data")
        if APP_STATE.get("vv_plan_path"):
            self._path_var.set(APP_STATE["vv_plan_path"])
        if existing_data:
            n_label = len(getattr(existing_data, "label_prds", []))
            n_ifu   = len(getattr(existing_data, "ifu_prds",   []))
            msg = (f"✔  Loaded: {getattr(existing_data, 'doc_title', 'V&V Plan')}\n"
                   f"   {n_label} Label PRD(s)  ·  {n_ifu} IFU PRD(s)")
            self._status_var.set(msg)
            self._status_lbl.configure(fg=COLORS["success"])
            self._load_btn.configure(text="Reload V&V Plan", bg=COLORS["accent2"])
            self._plan_data = existing_data
            self._populate_table(existing_data)

    def _center(self):
        self.update_idletasks()
        w, h = 860, 600
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.minsize(640, 420)

    # ── UI construction ───────────────────────────────────────────────────

    def _build_ui(self):
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)

        # ── Header ────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=COLORS["header_bg"], padx=20, pady=16)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="V&V Plan",
                 bg=COLORS["header_bg"], fg="white",
                 font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(hdr,
                 text="Load the V&V Plan document — requirements are shared automatically "
                      "with Label Verification and IFU Verification",
                 bg=COLORS["header_bg"], fg="#BFDBFE",
                 font=FONT_SMALL).pack(anchor="w", pady=(2, 0))

        # ── Body ──────────────────────────────────────────────────────────
        body = tk.Frame(self, bg=COLORS["bg_card"])
        body.grid(row=1, column=0, sticky="nsew")
        body.rowconfigure(3, weight=1)
        body.columnconfigure(0, weight=1)

        # Upload row
        upload_frame = tk.Frame(body, bg=COLORS["bg_card"], padx=24, pady=16)
        upload_frame.grid(row=0, column=0, sticky="ew")
        upload_frame.columnconfigure(1, weight=1)

        tk.Label(upload_frame, text="V&V Plan Document (.docx)",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                 font=FONT_SMALL).grid(row=0, column=0, columnspan=3,
                                       sticky="w", pady=(0, 4))
        tk.Entry(upload_frame, textvariable=self._path_var, state="readonly",
                 bg=COLORS["bg_input"], fg=COLORS["text"],
                 readonlybackground=COLORS["bg_input"],
                 relief="flat", font=FONT_MONO,
                 highlightthickness=1,
                 highlightbackground=COLORS["border"]).grid(
            row=1, column=0, sticky="ew", padx=(0, 6))
        tk.Button(upload_frame, text="Browse…",
                  command=self._browse,
                  bg=COLORS["bg_card"], fg=COLORS["text"],
                  activebackground=COLORS["border"],
                  relief="flat", cursor="hand2",
                  font=FONT_SMALL, padx=10, pady=4, bd=0,
                  highlightthickness=1,
                  highlightbackground=COLORS["border"]).grid(
            row=1, column=1, sticky="w")

        # Status label
        self._status_lbl = tk.Label(
            upload_frame, textvariable=self._status_var,
            bg=COLORS["bg_card"], fg=COLORS["text_muted"],
            font=FONT_SMALL, anchor="w",
            wraplength=700, justify="left")
        self._status_lbl.grid(row=2, column=0, columnspan=2,
                               sticky="w", pady=(8, 0))

        # Buttons row
        btn_row = tk.Frame(body, bg=COLORS["bg_card"])
        btn_row.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 16))

        self._load_btn = tk.Button(
            btn_row, text="Load V&V Plan",
            command=self._run,
            bg=COLORS["accent"], fg="white",
            activebackground=_darken(COLORS["accent"]),
            activeforeground="white",
            relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"),
            padx=16, pady=8, bd=0)
        self._load_btn.pack(side="left")

        self._export_btn = tk.Button(
            btn_row, text="Export to Excel",
            command=self._export_excel,
            state="disabled",
            bg=COLORS["bg_card"], fg=COLORS["text"],
            activebackground=COLORS["border"],
            relief="flat", cursor="hand2",
            font=FONT_SMALL, padx=14, pady=8, bd=0,
            highlightthickness=1,
            highlightbackground=COLORS["border"])
        self._export_btn.pack(side="left", padx=(10, 0))

        # Divider
        tk.Frame(body, bg=COLORS["border"], height=1).grid(
            row=2, column=0, sticky="ew")

        # ── Requirements table ────────────────────────────────────────────
        tbl_frame = tk.Frame(body, bg=COLORS["bg"])
        tbl_frame.grid(row=3, column=0, sticky="nsew")
        tbl_frame.rowconfigure(0, weight=1)
        tbl_frame.columnconfigure(0, weight=1)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("VVPlan.Treeview",
                        background=COLORS["bg"],
                        foreground=COLORS["text"],
                        fieldbackground=COLORS["bg"],
                        rowheight=26,
                        font=FONT_SMALL)
        style.configure("VVPlan.Treeview.Heading",
                        background=COLORS["bg_card"],
                        foreground=COLORS["text_muted"],
                        font=("Segoe UI", 8, "bold"),
                        relief="flat")
        style.map("VVPlan.Treeview",
                  background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "white")])

        self._tree = ttk.Treeview(
            tbl_frame,
            columns=self._COL_IDS,
            show="headings",
            selectmode="browse",
            style="VVPlan.Treeview")

        for col_id, hdr_txt, width in zip(
                self._COL_IDS, self._COL_HDRS, self._COL_W):
            self._tree.heading(col_id, text=hdr_txt,
                               command=lambda c=col_id: self._sort_col(c))
            stretch = col_id == "requirement_text"
            self._tree.column(col_id, width=width, minwidth=40,
                               anchor="w", stretch=stretch)

        self._tree.tag_configure("label_row",
                                 background=COLORS["bg"],
                                 foreground=COLORS["text"])
        self._tree.tag_configure("ifu_row",
                                 background="#EFF6FF",
                                 foreground=COLORS["text"])

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical",
                             command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        # Empty-state label (shown when table has no data)
        self._empty_lbl = tk.Label(
            tbl_frame,
            text="Load a V&V Plan to view extracted requirements here.",
            bg=COLORS["bg"], fg=COLORS["text_muted"],
            font=FONT_SMALL)
        self._empty_lbl.place(relx=0.5, rely=0.5, anchor="center")

    # ── Table helpers ─────────────────────────────────────────────────────

    def _populate_table(self, data):
        """Fill the treeview with all label and IFU PRDs."""
        self._tree.delete(*self._tree.get_children())
        all_prds = (
            [(p, "Label") for p in getattr(data, "label_prds", [])] +
            [(p, "IFU")   for p in getattr(data, "ifu_prds",   [])]
        )
        if not all_prds:
            self._empty_lbl.place(relx=0.5, rely=0.5, anchor="center")
            self._export_btn.configure(state="disabled")
            return
        self._empty_lbl.place_forget()
        for prd, prd_type in all_prds:
            tag = "label_row" if prd_type == "Label" else "ifu_row"
            req = getattr(prd, "requirement_text", "") or ""
            req_display = req.replace("\n", " ").strip()
            self._tree.insert("", "end", values=(
                getattr(prd, "prd_id",    "—"),
                prd_type,
                getattr(prd, "vv_method", "") or "—",
                req_display,
            ), tags=(tag,))
        self._export_btn.configure(state="normal")

    def _sort_col(self, col: str):
        """Toggle-sort the treeview by column."""
        rows = [(self._tree.set(iid, col), iid)
                for iid in self._tree.get_children("")]
        reverse = getattr(self, f"_sort_rev_{col}", False)
        rows.sort(key=lambda x: x[0].lower(), reverse=reverse)
        for idx, (_, iid) in enumerate(rows):
            self._tree.move(iid, "", idx)
        setattr(self, f"_sort_rev_{col}", not reverse)

    # ── File actions ──────────────────────────────────────────────────────

    def _browse(self):
        from tkinter import filedialog
        p = filedialog.askopenfilename(
            parent=self,
            title="Select V&V Plan Document",
            filetypes=[("Word Documents", "*.docx"), ("All files", "*.*")])
        if p:
            self._path_var.set(p)
            self._status_var.set("")

    def _run(self):
        path = self._path_var.get().strip()
        if not path:
            self._status_var.set("Please browse for the V&V Plan .docx file first.")
            self._status_lbl.configure(fg=COLORS["danger"])
            return
        self._load_btn.configure(state="disabled", text="Parsing…")
        self._export_btn.configure(state="disabled")
        self._status_var.set("Parsing V&V Plan…")
        self._status_lbl.configure(fg=COLORS["text_muted"])
        threading.Thread(target=self._parse_thread, args=(path,),
                         daemon=True).start()

    def _parse_thread(self, path: str):
        try:
            import sys as _sys
            if _HERE not in _sys.path:
                _sys.path.insert(0, _HERE)
            from vv_plan_parser import parse_vv_plan
            data = parse_vv_plan(path)
            APP_STATE["vv_plan_path"] = path
            APP_STATE["vv_plan_data"] = data
            n_label = len(getattr(data, "label_prds", []))
            n_ifu   = len(getattr(data, "ifu_prds",   []))
            msg = (f"✔  Loaded: {getattr(data,'doc_title','V&V Plan')}\n"
                   f"   {n_label} Label PRD(s)  ·  {n_ifu} IFU PRD(s)")
            self.after(0, self._set_done, msg, data)
            fire_event("vv_plan_ready", data)
        except Exception as e:
            logger.exception("V&V Plan parse failed")
            self.after(0, self._set_error, str(e))

    def _set_done(self, msg: str, data):
        self._plan_data = data
        self._status_var.set(msg)
        self._status_lbl.configure(fg=COLORS["success"])
        self._load_btn.configure(state="normal", text="Reload V&V Plan",
                                 bg=COLORS["accent2"])
        self._populate_table(data)

    def _set_error(self, msg: str):
        self._status_var.set(f"✗  Error: {msg}")
        self._status_lbl.configure(fg=COLORS["danger"])
        self._load_btn.configure(state="normal", text="Load V&V Plan",
                                 bg=COLORS["accent"])

    # ── Excel export ──────────────────────────────────────────────────────

    def _export_excel(self):
        if not self._plan_data:
            return
        from tkinter import filedialog
        import os
        default = (
            f"VVPlan_Requirements_"
            f"{getattr(self._plan_data,'doc_title','export').replace(' ','_')[:40]}.xlsx")
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Requirements as Excel",
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side)
            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "Requirements"

            # ── Metadata rows ──────────────────────────────────────────
            d  = self._plan_data
            meta = [
                ("Document Title",  getattr(d, "doc_title",   "")),
                ("ER Number",       getattr(d, "er_number",   "")),
                ("Requirement Doc", getattr(d, "req_doc",     "")),
                ("Sample Size",     getattr(d, "sample_size", "")),
            ]
            hdr_fill  = PatternFill("solid", fgColor="1E40AF")
            meta_fill = PatternFill("solid", fgColor="DBEAFE")
            bold_w    = Font(bold=True, color="FFFFFF")
            bold_b    = Font(bold=True, color="1E40AF")
            thin      = Side(style="thin", color="CBD5E1")
            border    = Border(left=thin, right=thin,
                               top=thin,  bottom=thin)

            for label, value in meta:
                ws.append([label, value])
                ws.cell(ws.max_row, 1).font  = bold_b
                ws.cell(ws.max_row, 1).fill  = meta_fill
                ws.cell(ws.max_row, 1).border = border
                ws.cell(ws.max_row, 2).border = border

            ws.append([])   # blank spacer

            # ── Column headers ─────────────────────────────────────────
            headers = ["PRD ID", "Type", "V&V Method",
                       "Requirement Text", "Symbols / References"]
            ws.append(headers)
            hdr_row = ws.max_row
            for col, _ in enumerate(headers, start=1):
                cell       = ws.cell(hdr_row, col)
                cell.font  = bold_w
                cell.fill  = hdr_fill
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)

            # ── Data rows ──────────────────────────────────────────────
            label_fill = PatternFill("solid", fgColor="F8FAFF")
            ifu_fill   = PatternFill("solid", fgColor="EFF6FF")
            center     = Alignment(horizontal="center", vertical="top")
            top_left   = Alignment(vertical="top", wrap_text=True)

            all_prds = (
                [(p, "Label") for p in getattr(d, "label_prds", [])] +
                [(p, "IFU")   for p in getattr(d, "ifu_prds",   [])]
            )
            for prd, prd_type in all_prds:
                syms = ", ".join(getattr(prd, "symbols", []) or [])
                ws.append([
                    getattr(prd, "prd_id",           ""),
                    prd_type,
                    getattr(prd, "vv_method",        "") or "",
                    getattr(prd, "requirement_text", "") or "",
                    syms,
                ])
                row_idx = ws.max_row
                fill    = label_fill if prd_type == "Label" else ifu_fill
                for col in range(1, 6):
                    cell        = ws.cell(row_idx, col)
                    cell.fill   = fill
                    cell.border = border
                    cell.alignment = (
                        center if col in (1, 2, 3) else top_left)

            # ── Column widths ──────────────────────────────────────────
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 70
            ws.column_dimensions["E"].width = 30
            ws.freeze_panes = ws.cell(hdr_row + 1, 1)

            wb.save(path)
            messagebox.showinfo(
                "Export Complete",
                f"Requirements exported to:\n{os.path.basename(path)}",
                parent=self)
        except Exception as e:
            logger.exception("Excel export failed")
            messagebox.showerror(
                "Export Failed", f"Could not write Excel file:\n{e}",
                parent=self)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN LAUNCHER — EVO MDD Verification Tool
# ══════════════════════════════════════════════════════════════════════════════

class LauncherApp:

    _TOOLS = [
        {
            "id":       "iso_symbol_extractor",
            "title":    "ISO Symbol Library",
            "subtitle": "Upload ISO 15223 standard PDF\nSymbols extracted automatically · One-time setup",
            "step":     "Step 1",
            "color":    "#16A34A",
            "icon":     "⬢",
            "module":   "iso_symbol_extractor",
            "btn":      "Open Symbol Library",
        },
        {
            "id":       "vv_plan",
            "title":    "V&V Plan",
            "subtitle": "Upload V&V Plan document (.docx)\nLoads requirements · One-time setup",
            "step":     "Step 2",
            "color":    "#7C3AED",
            "icon":     "📋",
            "module":   "vv_plan",
            "btn":      "Load V&V Plan",
        },
        {
            "id":       "label_verifier",
            "title":    "Label Verification",
            "subtitle": "Verify label symbols against\nISO 15223 & V&V Plan requirements",
            "step":     "Step 3",
            "color":    "#0B54A4",
            "icon":     "✦",
            "module":   "label_verifier",
            "btn":      "Open Label Verification",
        },
        {
            "id":       "ifu_verifier",
            "title":    "IFU / Addendum Verification",
            "subtitle": "Verify IFU document meets\nV&V Plan PRD requirements",
            "step":     "Step 4",
            "color":    "#E87722",
            "icon":     "📄",
            "module":   "ifu_verifier",
            "btn":      "Open IFU Verification",
        },
    ]

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("EVO MDD Verification Tool")
        self.root.configure(bg=COLORS["bg"])
        self.root.resizable(True, True)
        self._open_windows: dict = {}
        self._status_vars:  dict = {}
        self._build_ui()
        self._center()
        self._setup_automation_feedback()
        # Auto-check if symbol library already exists
        self._check_library_on_startup()

    def _center(self):
        self.root.update_idletasks()
        w, h = 980, 660
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.root.minsize(720, 500)

    def _check_library_on_startup(self):
        """Silently populate icon_library_path if library already exists."""
        if _library_is_built():
            APP_STATE["icon_library_path"] = _ICON_LIBRARY_DIR
            n = len([f for f in os.listdir(_ICON_LIBRARY_DIR)
                     if f.lower().endswith(".png")])
            self._set_tool_status("iso_symbol_extractor",
                                  f"✔  {n} symbols ready", COLORS["success"])
            self._set_tool_status("label_verifier",
                                  "⚡ Symbol library auto-loaded", COLORS["accent"])

    # ── Automation feedback ────────────────────────────────────────────────

    def _setup_automation_feedback(self):
        register_callback("icon_library_ready", self._on_icon_library_ready)
        register_callback("symbols_ready",      self._on_symbols_ready)
        register_callback("vv_plan_ready",      self._on_vv_plan_ready)

    def _on_icon_library_ready(self, path: str):
        n = len([f for f in os.listdir(path) if f.lower().endswith(".png")]) \
            if os.path.isdir(path) else 0
        self.root.after(0, self._set_tool_status, "iso_symbol_extractor",
                        f"✔  {n} symbols saved", COLORS["success"])
        self.root.after(0, self._set_tool_status, "label_verifier",
                        "⚡ Symbol library ready", COLORS["accent"])
        self.root.after(0, self._set_tool_status, "ifu_verifier",
                        "⚡ Symbol library ready", COLORS["accent"])

    def _on_symbols_ready(self, summary: list):
        needed = [s["symbol"] for s in summary if s.get("gtin_count", 0) > 0]
        if needed:
            self.root.after(0, self._set_tool_status, "label_verifier",
                            f"⚡ {', '.join(needed)} symbol(s) from V&V Plan",
                            COLORS["accent"])

    def _on_vv_plan_ready(self, data):
        n_label = len(getattr(data, "label_prds", []))
        n_ifu   = len(getattr(data, "ifu_prds",   []))
        self.root.after(0, self._set_tool_status, "vv_plan",
                        f"✔  {n_label} label · {n_ifu} IFU PRDs", COLORS["success"])
        self.root.after(0, self._set_tool_status, "label_verifier",
                        f"⚡ V&V Plan loaded — {n_label} PRD(s)", COLORS["accent"])
        self.root.after(0, self._set_tool_status, "ifu_verifier",
                        f"⚡ V&V Plan loaded — {n_ifu} IFU PRD(s)", COLORS["accent"])

    def _set_tool_status(self, tool_id: str, msg: str, color: str):
        entry = self._status_vars.get(tool_id)
        if entry:
            entry["text"].set(msg)
            entry["label"].configure(fg=color)

    # ── UI construction ────────────────────────────────────────────────────

    def _build_ui(self):
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=0)
        self.root.columnconfigure(1, weight=1)
        self._build_sidebar()
        self._build_main_area()

    def _build_sidebar(self):
        sidebar = tk.Frame(self.root, bg=COLORS["sidebar"], width=240)
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)

        # Brand block
        brand = tk.Frame(sidebar, bg=COLORS["header_bg"], pady=20, padx=18)
        brand.pack(fill="x")
        tk.Label(brand, text="EVO MDD", bg=COLORS["header_bg"],
                 fg="white", font=("Segoe UI", 16, "bold")).pack(anchor="w")
        tk.Label(brand, text="Verification Tool",
                 bg=COLORS["header_bg"], fg="#BFDBFE",
                 font=("Segoe UI", 9)).pack(anchor="w")
        tk.Label(brand, text="Philips Respironics · Systems V&V",
                 bg=COLORS["header_bg"], fg="#93C5FD",
                 font=FONT_SMALL).pack(anchor="w", pady=(4, 0))

        # Nav section
        tk.Label(sidebar, text="WORKFLOW", bg=COLORS["sidebar"],
                 fg="#7B9CC8", font=("Segoe UI", 7, "bold")).pack(
            anchor="w", padx=18, pady=(18, 4))

        for tool in self._TOOLS:
            self._make_nav_item(sidebar, tool)

        # Auto-link status card
        tk.Label(sidebar, text="AUTO-LINKS", bg=COLORS["sidebar"],
                 fg="#7B9CC8", font=("Segoe UI", 7, "bold")).pack(
            anchor="w", padx=18, pady=(22, 4))

        auto_card = tk.Frame(sidebar, bg="#0B3A7A",
                             highlightthickness=1,
                             highlightbackground="#1E5FAD")
        auto_card.pack(fill="x", padx=12, pady=2)
        inner = tk.Frame(auto_card, bg="#0B3A7A", padx=12, pady=10)
        inner.pack(fill="x")
        tk.Label(inner, text="✔  Automation active",
                 bg="#0B3A7A", fg="#4ADE80",
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Label(inner,
                 text="Step 1 → Steps 3 & 4  icon library\nStep 2 → Steps 3 & 4  requirements",
                 bg="#0B3A7A", fg="#93C5FD",
                 font=FONT_SMALL, justify="left").pack(anchor="w", pady=(4, 0))

        # Version footer
        tk.Label(sidebar, text="v2.0  ·  EVO MDD",
                 bg=COLORS["sidebar"], fg="#3D5A80",
                 font=FONT_SMALL).pack(side="bottom", pady=10)

    def _make_nav_item(self, parent, tool: dict):
        item = tk.Frame(parent, bg=COLORS["sidebar"], cursor="hand2")
        item.pack(fill="x", padx=6, pady=1)

        bar = tk.Frame(item, bg=COLORS["sidebar"], width=4)
        bar.pack(side="left", fill="y")

        content = tk.Frame(item, bg=COLORS["sidebar"], padx=10, pady=9)
        content.pack(side="left", fill="both", expand=True)

        step_lbl = tk.Label(content, text=tool["step"],
                             bg=COLORS["sidebar"], fg="#7B9CC8",
                             font=("Segoe UI", 7, "bold"))
        step_lbl.pack(anchor="w")

        title_lbl = tk.Label(content,
                              text=tool["title"].replace("\n", " "),
                              bg=COLORS["sidebar"], fg="white",
                              font=("Segoe UI", 9, "bold"), anchor="w")
        title_lbl.pack(anchor="w")

        status_var = tk.StringVar(value="")
        status_lbl = tk.Label(content, textvariable=status_var,
                               bg=COLORS["sidebar"], fg="#93C5FD",
                               font=FONT_SMALL, anchor="w")
        status_lbl.pack(anchor="w")
        self._status_vars[tool["id"]] = {"text": status_var, "label": status_lbl}

        color      = tool["color"]
        all_widgets = [item, bar, content, step_lbl, title_lbl, status_lbl]

        def on_enter(e, ws=all_widgets, b=bar, c=color):
            for w in ws: w.configure(bg=COLORS["sidebar_hover"])
            b.configure(bg=c, width=4)

        def on_leave(e, ws=all_widgets, b=bar):
            for w in ws: w.configure(bg=COLORS["sidebar"])
            b.configure(bg=COLORS["sidebar"], width=4)

        for w in all_widgets:
            w.bind("<Enter>",    on_enter)
            w.bind("<Leave>",    on_leave)
            w.bind("<Button-1>", lambda e, m=tool["module"]: self._open_tool(m))

    def _build_main_area(self):
        main = tk.Frame(self.root, bg=COLORS["bg"])
        main.grid(row=0, column=1, sticky="nsew")
        main.rowconfigure(1, weight=1)
        main.columnconfigure(0, weight=1)

        # Top bar
        topbar = tk.Frame(main, bg=COLORS["bg_card"],
                          highlightthickness=1,
                          highlightbackground=COLORS["border"],
                          padx=28, pady=16)
        topbar.grid(row=0, column=0, sticky="ew")
        tk.Label(topbar, text="EVO MDD Verification Tool",
                 bg=COLORS["bg_card"], fg=COLORS["accent"],
                 font=("Segoe UI", 15, "bold")).pack(anchor="w")
        tk.Label(topbar,
                 text="Automated medical symbol verification  ·  "
                      "Trilogy EVO / Helix  ·  Philips Respironics",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                 font=FONT_BODY).pack(anchor="w", pady=(3, 0))

        # 2×2 card grid (4 tools: ISO · V&V Plan / Label · IFU)
        cards_area = tk.Frame(main, bg=COLORS["bg"])
        cards_area.grid(row=1, column=0, sticky="nsew", padx=20, pady=16)
        cards_area.columnconfigure(0, weight=1)
        cards_area.columnconfigure(1, weight=1)
        cards_area.rowconfigure(0, weight=1)
        cards_area.rowconfigure(1, weight=1)

        for i, tool in enumerate(self._TOOLS):
            self._make_tool_card(cards_area, tool, i // 2, i % 2)

        # Info footer
        info = tk.Frame(main, bg=COLORS["bg_card"],
                        highlightthickness=1,
                        highlightbackground=COLORS["border"],
                        padx=24, pady=8)
        info.grid(row=2, column=0, sticky="ew")
        tk.Label(info,
                 text="Steps 1 & 2 are one-time setup  ·  "
                      "Step 3: Label Verification uses the ISO library + V&V Plan  ·  "
                      "Step 4: IFU / Addendum Verification uses the V&V Plan",
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                 font=FONT_SMALL).pack(anchor="w")

    def _make_tool_card(self, parent, tool: dict, row: int, col: int):
        color = tool["color"]
        pad_l = 0 if col == 0 else 6
        pad_r = 0 if col == len(self._TOOLS) - 1 else 6

        card = tk.Frame(parent, bg=COLORS["bg_card"],
                        highlightthickness=1,
                        highlightbackground=COLORS["border"],
                        cursor="hand2")
        card.grid(row=row, column=col, sticky="nsew",
                  padx=(pad_l, pad_r), pady=4)
        card.columnconfigure(0, weight=1)

        # Top accent strip
        tk.Frame(card, bg=color, height=4).pack(fill="x")

        body = tk.Frame(card, bg=COLORS["bg_card"], padx=20, pady=18)
        body.pack(fill="both", expand=True)

        # Step badge
        badge = tk.Label(body, text=tool["step"], bg=color, fg="white",
                         font=("Segoe UI", 7, "bold"), padx=8, pady=3)
        badge.pack(anchor="w", pady=(0, 10))

        # Icon
        tk.Label(body, text=tool["icon"],
                 bg=COLORS["bg_card"], fg=color,
                 font=("Segoe UI", 24)).pack(anchor="w")

        # Title (multi-line friendly)
        for line in tool["title"].split("\n"):
            tk.Label(body, text=line,
                     bg=COLORS["bg_card"], fg=COLORS["text"],
                     font=("Segoe UI", 11, "bold"), anchor="w").pack(anchor="w")

        # Subtitle lines
        tk.Frame(body, bg=COLORS["border"], height=1).pack(fill="x", pady=(8, 8))
        for line in tool["subtitle"].split("\n"):
            tk.Label(body, text=line,
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                     font=FONT_SMALL, anchor="w").pack(anchor="w")

        # Open button
        btn = tk.Button(
            body, text=tool["btn"],
            command=lambda m=tool["module"]: self._open_tool(m),
            bg=color, fg="white",
            activebackground=_darken(color), activeforeground="white",
            relief="flat", cursor="hand2",
            font=("Segoe UI", 9, "bold"),
            padx=16, pady=8, bd=0,
        )
        btn.pack(anchor="w", pady=(14, 0))

        # Hover effect on card border
        def _enter(e, c=card, col=color): c.configure(highlightbackground=col)
        def _leave(e, c=card):            c.configure(highlightbackground=COLORS["border"])
        for w in (card, body):
            w.bind("<Enter>", _enter)
            w.bind("<Leave>", _leave)

    # ── Open tool ─────────────────────────────────────────────────────────

    def _open_tool(self, module_name: str):
        # Prevent duplicate windows (each module gets at most one window)
        existing = self._open_windows.get(module_name)
        if existing and existing.winfo_exists():
            self._focus_window(existing)
            return

        try:
            if module_name == "iso_symbol_extractor":
                win = ISOLibraryUploadWindow(self.root)
                self._open_windows[module_name] = win
                self._focus_window(win)
                return

            if module_name == "vv_plan":
                win = VVPlanWindow(self.root)
                self._open_windows[module_name] = win
                self._focus_window(win)
                return

            if module_name == "label_verifier":
                from label_verifier.gui.main_gui import MainGUI
                win = tk.Toplevel(self.root)
                MainGUI(win, mode="label")
                self._open_windows[module_name] = win
                self._focus_window(win)
                return

            if module_name == "ifu_verifier":
                from ifu_verifier_window import IFUVerifierWindow
                plan_data = APP_STATE.get("vv_plan_data")
                ifu_prds  = (list(getattr(plan_data, "ifu_prds", []))
                             if plan_data else [])
                if not ifu_prds and not plan_data:
                    messagebox.showinfo(
                        "Load V&V Plan First",
                        "Please load a V&V Plan (Step 2) before opening "
                        "IFU Verification.\n\n"
                        "The plan provides the IFU requirements to verify.",
                        parent=self.root)
                    return
                win = IFUVerifierWindow(self.root,
                                        ifu_prds=ifu_prds,
                                        plan_data=plan_data)
                self._open_windows[module_name] = win
                self._focus_window(win)
                return

        except ImportError as exc:
            messagebox.showerror(
                "Module Not Found",
                f"Could not load '{module_name}'.\n\n{exc}\n\n"
                "Make sure all files are in the same directory as app.py.",
                parent=self.root,
            )
        except Exception as exc:
            logger.exception("Failed to open tool: %s", module_name)
            messagebox.showerror("Error", f"Failed to open tool:\n{exc}",
                                 parent=self.root)

    @staticmethod
    def _focus_window(win: tk.Toplevel):
        win.lift()
        win.focus_force()
        win.attributes("-topmost", True)
        win.after(200, lambda: win.attributes("-topmost", False))


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_app():
    root = tk.Tk()
    LauncherApp(root)
    root.mainloop()


def _log_exception(exc_type, exc_value, exc_tb):
    logger.exception("Unhandled exception", exc_info=(exc_type, exc_value, exc_tb))


def _thread_exception_handler(args):
    exc_type  = getattr(args, "exc_type",      None)
    exc_value = getattr(args, "exc_value",     None)
    exc_tb    = getattr(args, "exc_traceback", None)
    if exc_type and exc_value and exc_tb:
        _log_exception(exc_type, exc_value, exc_tb)


if __name__ == "__main__":
    sys.excepthook = _log_exception
    try:
        threading.excepthook = _thread_exception_handler
    except AttributeError:
        pass
    try:
        run_app()
    except Exception:
        logger.exception("Exception in run_app")
        sys.exit(1)