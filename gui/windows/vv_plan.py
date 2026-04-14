import os
import threading
import tkinter as tk
from tkinter import ttk, messagebox
import logging

from gui.theme import (
    COLORS, FONT_SMALL, FONT_MONO, darken, contrast_text,
)
from core.state_manager import APP_STATE, fire_event

logger = logging.getLogger(__name__)
_HERE = os.path.dirname(os.path.abspath(__file__))

VV_LIGHT_COLORS = dict(COLORS)
VV_DARK_COLORS = {
    **VV_LIGHT_COLORS,
    "bg": "#0F172A",
    "bg_card": "#111827",
    "bg_input": "#0B1220",
    "accent": "#60A5FA",
    "text": "#E5E7EB",
    "text_muted": "#94A3B8",
    "border": "#334155",
    "header_bg": "#1D4ED8",
    "header_subtle": "#BFDBFE",
    "success": "#4ADE80",
    "danger": "#F87171",
}

class VVPlanWindow(tk.Toplevel):
    _COL_IDS  = ("prd_id", "type", "vv_method", "requirement_text")
    _COL_HDRS = ("PRD ID", "Type", "V&V Method", "Requirement Text")
    _COL_W    = (90, 60, 100, 500)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("V&V Plan — EVO MDD Verification Tool")
        self.resizable(True, True)
        self._path_var    = tk.StringVar()
        self._status_var  = tk.StringVar(value="")
        self._plan_data   = None
        self._theme_name = "light"
        self._status_color = "text_muted"
        self._load_btn_mode = "load"
        self._is_running = False
        self._render_ui()
        self._center()

        existing_data = APP_STATE.get("vv_plan_data")
        if APP_STATE.get("vv_plan_path"):
            self._path_var.set(APP_STATE["vv_plan_path"])
        if existing_data:
            n_label = len(getattr(existing_data, "label_prds", []))
            n_ifu   = len(getattr(existing_data, "ifu_prds",   []))
            msg = (f"✔  Loaded: {getattr(existing_data, 'doc_title', 'V&V Plan')}\n"
                   f"   {n_label} Label PRD(s)  ·  {n_ifu} IFU PRD(s)")
            self._status_var.set(msg)
            self._status_color = "success"
            self._load_btn_mode = "reload"
            self._apply_runtime_state()
            self._plan_data = existing_data
            self._populate_table(existing_data)

    def _colors(self) -> dict:
        return VV_LIGHT_COLORS if self._theme_name == "light" else VV_DARK_COLORS

    def _theme_button_label(self) -> str:
        return "☀ Light" if self._theme_name == "dark" else "☾ Dark"

    def _toggle_theme(self):
        self._theme_name = "dark" if self._theme_name == "light" else "light"
        self._apply_theme_in_place()

    def _render_ui(self):
        self.configure(bg=self._colors()["bg_card"])
        for child in self.winfo_children():
            child.destroy()
        self._build_ui()
        self._apply_runtime_state()
        if self._plan_data:
            self._populate_table(self._plan_data)

    def _apply_theme_in_place(self):
        c = self._colors()
        self.configure(bg=c["bg_card"])
        self._hdr.configure(bg=c["header_bg"])
        self._theme_btn.configure(
            text=self._theme_button_label(),
            bg=c["header_bg"],
            fg="white",
            activebackground=darken(c["header_bg"]),
            activeforeground="white",
        )
        self._title_lbl.configure(bg=c["header_bg"], fg="white")
        self._subtitle_lbl.configure(bg=c["header_bg"], fg=c["header_subtle"])
        self._body.configure(bg=c["bg_card"])
        self._upload_frame.configure(bg=c["bg_card"])
        self._doc_lbl.configure(bg=c["bg_card"], fg=c["text_muted"])
        self._path_entry.configure(
            bg=c["bg_input"],
            fg=c["text"],
            readonlybackground=c["bg_input"],
            highlightbackground=c["border"],
            insertbackground=c["text"],
        )
        self._browse_btn.configure(
            bg=c["bg_card"],
            fg=c["text"],
            activebackground=c["border"],
            activeforeground=c["text"],
            highlightbackground=c["border"],
        )
        self._btn_row.configure(bg=c["bg_card"])
        self._divider.configure(bg=c["border"])
        self._tbl_frame.configure(bg=c["bg"])
        self._empty_lbl.configure(bg=c["bg"], fg=c["text_muted"])
        self._style_treeview()
        self._apply_runtime_state()

    def _apply_runtime_state(self):
        c = self._colors()
        self._status_lbl.configure(fg=c.get(self._status_color, self._status_color))
        self._style_treeview()

        if self._load_btn_mode == "reload":
            bg = c["accent2"]
            fg = contrast_text(bg)
            self._load_btn.configure(
                state="normal",
                text="Reload V&V Plan",
                bg=bg,
                fg=fg,
                activebackground=darken(bg),
                activeforeground=fg,
            )
        else:
            bg = c["accent"]
            fg = contrast_text(bg)
            self._load_btn.configure(
                state="disabled" if self._is_running else "normal",
                text="Parsing…" if self._is_running else "Load V&V Plan",
                bg=bg,
                fg=fg,
                activebackground=darken(bg),
                activeforeground=fg,
            )

        self._export_btn.configure(
            bg=c["accent2"] if self._plan_data else c["bg_card"],
            fg=contrast_text(c["accent2"]) if self._plan_data else c["text"],
            activebackground=darken(c["accent2"]) if self._plan_data else c["border"],
            activeforeground=contrast_text(c["accent2"]) if self._plan_data else c["text"],
            highlightbackground=c["accent2"] if self._plan_data else c["border"],
            highlightcolor=c["accent2"] if self._plan_data else c["border"],
            state="normal" if self._plan_data else "disabled",
        )

    def _style_treeview(self):
        c = self._colors()
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(
            "VVPlan.Treeview",
            background=c["bg"],
            foreground=c["text"],
            fieldbackground=c["bg"],
            rowheight=26,
            font=FONT_SMALL,
            bordercolor=c["border"],
            lightcolor=c["border"],
            darkcolor=c["border"],
        )
        style.configure(
            "VVPlan.Treeview.Heading",
            background=c["bg_card"],
            foreground=c["text_muted"],
            font=("Segoe UI", 8, "bold"),
            relief="flat",
            bordercolor=c["border"],
            lightcolor=c["border"],
            darkcolor=c["border"],
        )
        style.map("VVPlan.Treeview", background=[("selected", c["accent"])], foreground=[("selected", contrast_text(c["accent"]))])
        if hasattr(self, "_tree"):
            self._tree.tag_configure("label_row", background=c["bg"], foreground=c["text"])
            ifu_bg = "#172554" if self._theme_name == "dark" else "#EFF6FF"
            self._tree.tag_configure("ifu_row", background=ifu_bg, foreground=c["text"])

    def _center(self):
        self.update_idletasks()
        w, h = 860, 600
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.minsize(640, 420)

    def _build_ui(self):
        c = self._colors()
        self.rowconfigure(1, weight=1)
        self.columnconfigure(0, weight=1)

        hdr = tk.Frame(self, bg=c["header_bg"], padx=20, pady=16)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(0, weight=1)
        self._hdr = hdr
        self._theme_btn = tk.Button(
            hdr,
            text=self._theme_button_label(),
            command=self._toggle_theme,
            bg=c["header_bg"],
            fg="white",
            activebackground=darken(c["header_bg"]),
            activeforeground="white",
            relief="flat",
            cursor="hand2",
            font=FONT_SMALL,
            padx=10,
            pady=4,
            bd=0,
        )
        self._theme_btn.pack(anchor="e")
        self._title_lbl = tk.Label(hdr, text="V&V Plan", bg=c["header_bg"], fg="white", font=("Segoe UI", 13, "bold"))
        self._title_lbl.pack(anchor="w")
        self._subtitle_lbl = tk.Label(hdr, text="Load the V&V Plan document — requirements are shared automatically with Label Verification and IFU Verification", bg=c["header_bg"], fg=c["header_subtle"], font=FONT_SMALL)
        self._subtitle_lbl.pack(anchor="w", pady=(2, 0))

        body = tk.Frame(self, bg=c["bg_card"])
        body.grid(row=1, column=0, sticky="nsew")
        body.rowconfigure(3, weight=1)
        body.columnconfigure(0, weight=1)
        self._body = body

        upload_frame = tk.Frame(body, bg=c["bg_card"], padx=24, pady=16)
        upload_frame.grid(row=0, column=0, sticky="ew")
        upload_frame.columnconfigure(1, weight=1)
        self._upload_frame = upload_frame

        self._doc_lbl = tk.Label(upload_frame, text="V&V Plan Document (.docx)", bg=c["bg_card"], fg=c["text_muted"], font=FONT_SMALL)
        self._doc_lbl.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 4))
        self._path_entry = tk.Entry(upload_frame, textvariable=self._path_var, state="readonly", bg=c["bg_input"], fg=c["text"], readonlybackground=c["bg_input"], relief="flat", font=FONT_MONO, highlightthickness=1, highlightbackground=c["border"], insertbackground=c["text"])
        self._path_entry.grid(row=1, column=0, sticky="ew", padx=(0, 6))
        self._browse_btn = tk.Button(upload_frame, text="Browse…", command=self._browse, bg=c["bg_card"], fg=c["text"], activebackground=c["border"], activeforeground=c["text"], relief="flat", cursor="hand2", font=FONT_SMALL, padx=10, pady=4, bd=0, highlightthickness=1, highlightbackground=c["border"])
        self._browse_btn.grid(row=1, column=1, sticky="w")

        self._status_lbl = tk.Label(upload_frame, textvariable=self._status_var, bg=c["bg_card"], fg=c["text_muted"], font=FONT_SMALL, anchor="w", wraplength=700, justify="left")
        self._status_lbl.grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 0))

        btn_row = tk.Frame(body, bg=c["bg_card"])
        btn_row.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 16))
        self._btn_row = btn_row

        self._load_btn = tk.Button(btn_row, text="Load V&V Plan", command=self._run, bg=c["accent"], fg=contrast_text(c["accent"]), activebackground=darken(c["accent"]), activeforeground=contrast_text(c["accent"]), relief="flat", cursor="hand2", font=("Segoe UI", 9, "bold"), padx=16, pady=8, bd=0)
        self._load_btn.pack(side="left")

        self._export_btn = tk.Button(btn_row, text="Export to Excel", command=self._export_excel, state="disabled", bg=c["bg_card"], fg=c["text"], activebackground=c["border"], activeforeground=c["text"], relief="flat", cursor="hand2", font=FONT_SMALL, padx=14, pady=8, bd=0, highlightthickness=1, highlightbackground=c["border"])
        self._export_btn.pack(side="left", padx=(10, 0))

        self._divider = tk.Frame(body, bg=c["border"], height=1)
        self._divider.grid(row=2, column=0, sticky="ew")

        tbl_frame = tk.Frame(body, bg=c["bg"])
        tbl_frame.grid(row=3, column=0, sticky="nsew")
        tbl_frame.rowconfigure(0, weight=1)
        tbl_frame.columnconfigure(0, weight=1)
        self._tbl_frame = tbl_frame

        self._style_treeview()

        self._tree = ttk.Treeview(tbl_frame, columns=self._COL_IDS, show="headings", selectmode="browse", style="VVPlan.Treeview")

        for col_id, hdr_txt, width in zip(self._COL_IDS, self._COL_HDRS, self._COL_W):
            self._tree.heading(col_id, text=hdr_txt, anchor="w", command=lambda c=col_id: self._sort_col(c))
            stretch = col_id == "requirement_text"
            self._tree.column(col_id, width=width, minwidth=40, anchor="w", stretch=stretch)

        self._tree.tag_configure("label_row", background=c["bg"], foreground=c["text"])
        self._tree.tag_configure("ifu_row", background="#172554" if self._theme_name == "dark" else "#EFF6FF", foreground=c["text"])

        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self._empty_lbl = tk.Label(tbl_frame, text="Load a V&V Plan to view extracted requirements here.", bg=c["bg"], fg=c["text_muted"], font=FONT_SMALL)
        self._empty_lbl.place(relx=0.5, rely=0.5, anchor="center")

    def _populate_table(self, data):
        try:
            self._tree.get_children()
        except tk.TclError:
            return
        self._tree.delete(*self._tree.get_children())
        all_prds = ([(p, "Label") for p in getattr(data, "label_prds", [])] + [(p, "IFU") for p in getattr(data, "ifu_prds", [])])
        if not all_prds:
            self._empty_lbl.place(relx=0.5, rely=0.5, anchor="center")
            self._export_btn.configure(state="disabled")
            return
        self._empty_lbl.place_forget()
        for prd, prd_type in all_prds:
            tag = "label_row" if prd_type == "Label" else "ifu_row"
            req = getattr(prd, "requirement_text", "") or ""
            req_display = req.replace("\n", " ").strip()
            self._tree.insert("", "end", values=(getattr(prd, "prd_id", "—"), prd_type, getattr(prd, "vv_method", "") or "—", req_display), tags=(tag,))
        self._export_btn.configure(state="normal")

    def _sort_col(self, col: str):
        if not self._tree_ready():
            return
        rows = [(self._tree.set(iid, col), iid) for iid in self._tree.get_children("")]
        reverse = getattr(self, f"_sort_rev_{col}", False)
        rows.sort(key=lambda x: x[0].lower(), reverse=reverse)
        for idx, (_, iid) in enumerate(rows):
            self._tree.move(iid, "", idx)
        setattr(self, f"_sort_rev_{col}", not reverse)

    def _browse(self):
        from tkinter import filedialog
        p = filedialog.askopenfilename(parent=self, title="Select V&V Plan Document", filetypes=[("Word Documents", "*.docx"), ("All files", "*.*")])
        if p:
            self._path_var.set(p)
            self._status_var.set("")
            self._status_color = "text_muted"
            self._status_lbl.configure(fg=self._colors()["text_muted"])

    def _run(self):
        path = self._path_var.get().strip()
        if not path:
            self._status_var.set("Please browse for the V&V Plan .docx file first.")
            self._status_color = "danger"
            self._status_lbl.configure(fg=self._colors()["danger"])
            return
        self._is_running = True
        self._load_btn_mode = "load"
        self._load_btn.configure(state="disabled", text="Parsing…")
        self._export_btn.configure(state="disabled")
        self._status_var.set("Parsing V&V Plan…")
        self._status_color = "text_muted"
        self._status_lbl.configure(fg=self._colors()["text_muted"])
        threading.Thread(target=self._parse_thread, args=(path,), daemon=True).start()

    def _parse_thread(self, path: str):
        try:
            import sys as _sys
            _root_dir = os.path.normpath(os.path.join(_HERE, "..", ".."))
            if _root_dir not in _sys.path:
                _sys.path.insert(0, _root_dir)
            from parsers.vv_plan_parser import parse_vv_plan
            data = parse_vv_plan(path)
            APP_STATE["vv_plan_path"] = path
            APP_STATE["vv_plan_data"] = data
            n_label = len(getattr(data, "label_prds", []))
            n_ifu   = len(getattr(data, "ifu_prds",   []))
            msg = (f"✔  Loaded: {getattr(data,'doc_title','V&V Plan')}\n"
                   f"   {n_label} Label PRD(s)  ·  {n_ifu} IFU PRD(s)")
            self.after(0, self._set_done, msg, data)
            fire_event("vv_plan_ready", data)
        except Exception as e:
            logger.exception("V&V Plan parse failed")
            self.after(0, self._set_error, str(e))

    def _set_done(self, msg: str, data):
        if not self.winfo_exists():
            return
        self._plan_data = data
        self._is_running = False
        self._load_btn_mode = "reload"
        self._status_var.set(msg)
        self._status_color = "success"
        self._apply_runtime_state()
        self._populate_table(data)

    def _set_error(self, msg: str):
        if not self.winfo_exists():
            return
        self._is_running = False
        self._load_btn_mode = "load"
        self._status_var.set(f"✗  Error: {msg}")
        self._status_color = "danger"
        self._apply_runtime_state()

    def _export_excel(self):
        if not self._plan_data:
            return
        from tkinter import filedialog
        default = f"VVPlan_Requirements_{getattr(self._plan_data,'doc_title','export').replace(' ','_')[:40]}.xlsx"
        path = filedialog.asksaveasfilename(parent=self, title="Save Requirements as Excel", defaultextension=".xlsx", initialfile=default, filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = "Requirements"
            
            d  = self._plan_data
            meta = [
                ("Document Title",  getattr(d, "doc_title",   "")),
                ("ER Number",       getattr(d, "er_number",   "")),
                ("Requirement Doc", getattr(d, "req_doc",     "")),
                ("Sample Size",     getattr(d, "sample_size", "")),
            ]
            hdr_fill  = PatternFill("solid", fgColor="1E40AF")
            meta_fill = PatternFill("solid", fgColor="DBEAFE")
            bold_w    = Font(bold=True, color="FFFFFF")
            bold_b    = Font(bold=True, color="1E40AF")
            thin      = Side(style="thin", color="CBD5E1")
            border    = Border(left=thin, right=thin, top=thin,  bottom=thin)

            for label, value in meta:
                ws.append([label, value])
                ws.cell(ws.max_row, 1).font  = bold_b
                ws.cell(ws.max_row, 1).fill  = meta_fill
                ws.cell(ws.max_row, 1).border = border
                ws.cell(ws.max_row, 2).border = border
            ws.append([])

            headers = ["PRD ID", "Type", "V&V Method", "Requirement Text", "Symbols / References"]
            ws.append(headers)
            hdr_row = ws.max_row
            for col, _ in enumerate(headers, start=1):
                cell       = ws.cell(hdr_row, col)
                cell.font  = bold_w
                cell.fill  = hdr_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            label_fill = PatternFill("solid", fgColor="F8FAFF")
            ifu_fill   = PatternFill("solid", fgColor="EFF6FF")
            center     = Alignment(horizontal="center", vertical="top")
            top_left   = Alignment(vertical="top", wrap_text=True)

            all_prds = ([(p, "Label") for p in getattr(d, "label_prds", [])] + [(p, "IFU") for p in getattr(d, "ifu_prds", [])])
            for prd, prd_type in all_prds:
                syms = ", ".join(getattr(prd, "symbols", []) or [])
                ws.append([getattr(prd, "prd_id", ""), prd_type, getattr(prd, "vv_method", "") or "", getattr(prd, "requirement_text", "") or "", syms])
                row_idx = ws.max_row
                fill    = label_fill if prd_type == "Label" else ifu_fill
                for col in range(1, 6):
                    cell        = ws.cell(row_idx, col)
                    cell.fill   = fill
                    cell.border = border
                    cell.alignment = center if col in (1, 2, 3) else top_left

            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 70
            ws.column_dimensions["E"].width = 30
            ws.freeze_panes = ws.cell(hdr_row + 1, 1)

            wb.save(path)
            messagebox.showinfo("Export Complete", f"Requirements exported to:\n{os.path.basename(path)}", parent=self)
        except Exception as e:
            logger.exception("Excel export failed")
            messagebox.showerror("Export Failed", f"Could not write Excel file:\n{e}", parent=self)