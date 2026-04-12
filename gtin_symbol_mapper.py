"""
GTIN Symbol Mapper — Window 1
================================
All core logic preserved from original firstlayers_edited.py.
Enhanced with:
  • Progressive pop-up progress windows for each processing step
  • Modern dark-themed UI
  • Detail log panel with color-coded messages
  • Results table with full PRD column display
"""

import os
import re
import csv
import json
import threading
from datetime import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Import shared styles from launcher
try:
    from app import (
        COLORS, FONT_TITLE, FONT_HEADING, FONT_BODY, FONT_MONO, FONT_SMALL,
        ProgressPopup, styled_button, styled_label, styled_entry, _lighten,
        APP_STATE, fire_event
    )
except ImportError:
    # Fallback if run standalone
    APP_STATE = None
    def fire_event(event, *args): pass
    COLORS = {
        # Philips light theme
        "bg":           "#F5F7FA",
        "bg_card":      "#FFFFFF",
        "bg_input":     "#FFFFFF",
        "accent":       "#0B54A4",
        "accent2":      "#00A0DC",
        "accent3":      "#E87722",
        "danger":       "#DC2626",
        "text":         "#1A1A2E",
        "text_muted":   "#6B7280",
        "border":       "#D1D5DB",
        "header_bg":    "#0B54A4",
        "success":      "#16A34A",
        "warn":         "#D97706",
        "log_bg":       "#F8FAFC",
        "log_text":     "#1E3A5F",
        "sidebar":      "#0B2B5C",
        "sidebar_hover":"#0E3870",
    }
    FONT_TITLE   = ("Segoe UI", 16, "bold")
    FONT_HEADING = ("Segoe UI", 11, "bold")
    FONT_BODY    = ("Segoe UI", 9)
    FONT_MONO    = ("Consolas", 9)
    FONT_SMALL   = ("Segoe UI", 8)

    def _lighten(hex_color):
        hex_color = hex_color.lstrip("#")
        r, g, b = [int(hex_color[i:i+2], 16) for i in (0, 2, 4)]
        return f"#{min(255,r+30):02x}{min(255,g+30):02x}{min(255,b+30):02x}"


# ══════════════════════════════════════════════════════════════════════════════
# CORE PROCESSING LOGIC (preserved exactly from original)
# ══════════════════════════════════════════════════════════════════════════════

SYMBOL_PATTERNS: dict[str, list[str]] = {
    "MD": [
        r"MD\s+symbol",
        r"'MD'\s*\)\s*symbol",
        r"medical\s+device\s+symbol",
        r"medical\s+device.{0,40}symbol",
        r"symbol\s+5\s*[.,]\s*7\s*[.,]\s*7",
        r"5\s*[.,]\s*7\s*[.,]\s*7",
        r"ISO\s+15223.{0,40}5\s*[.,]\s*7\s*[.,]\s*7",
    ],
    "UDI": [
        r"\bUDI\b",
        r"unique\s+device\s+identifier",
        r"6\.1\.4\s*[\(\[]?\s*b",
    ],
    "Date of Manufacture": [
        r"date\s+of\s+manufacture",
        r"date\s+of\s+mfg",
        r"manufacturing\s+date",
        r"6\.1\.4\s*[\(\[]?\s*a.?4",
    ],
    "IFU": [
        r"consult\s+instructions?\s+for\s+use",
        r"consult\s+IFU",
        r"IFU\s+symbol",
        r"symbol\s+5\s*[.,]\s*4\s*[.,]\s*3",
        r"5\s*[.,]\s*4\s*[.,]\s*3",
        r"ISO\s+15223.{0,40}5\s*[.,]\s*4\s*[.,]\s*3",
        r"6\.1\.3\b",
        r"6\.1\.5\b",
    ],
}

ALL_SYMBOLS = ["MD", "UDI", "Date of Manufacture", "IFU"]
PRD_ID_RE   = re.compile(r"\b([A-Z]+(?:_[A-Z0-9]+-?[A-Z0-9]*)?_PRD\d+)\b")


def cell_str(val) -> str:
    return "" if val is None else str(val).strip()


def matches_any(text: str, patterns: list[str]) -> bool:
    for p in patterns:
        try:
            if re.search(p, text, re.IGNORECASE):
                return True
        except re.error:
            if p.lower() in text.lower():
                return True
    return False


def prd_col(sym: str) -> str:
    return f"PRD for {sym}"


def build_pn_to_symbols(filepath: str, log, progress_cb=None):
    log(f"[Step 1] Reading: {os.path.basename(filepath)}", "INFO")
    if progress_cb:
        progress_cb("Loading symbol matrix workbook…")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    def find_row(label):
        for row in ws.iter_rows():
            if row[0].value and label.lower() in str(row[0].value).lower():
                return [c.value for c in row]
        return None

    if progress_cb:
        progress_cb("Scanning for required rows…")

    fgn_row = find_row("FGN")
    er_row  = find_row("PRD ER No")
    md_row  = find_row("MD Symbol missing")
    udi_row = find_row("UDI Symbol missing")
    dom_row = find_row("Date of manufacture Symbol missing")
    ifu_row = find_row("IFU Symbol missing")

    if not fgn_row:
        raise ValueError("Could not find 'FGN' row in File 1.")

    def normalise_er(raw: str) -> str:
        m = re.search(r"\d{5,}", raw)
        return f"ER{m.group()}" if m else raw.strip()

    pn_to_symbols: dict[str, list[str]] = {}
    pn_to_er: dict[str, str] = {}
    last_er = ""

    for col_idx in range(1, len(fgn_row)):
        pn = cell_str(fgn_row[col_idx])
        if not pn:
            continue
        if er_row:
            v = cell_str(er_row[col_idx])
            if v:
                last_er = normalise_er(v)
        pn_to_er[pn] = last_er

        syms = []
        if md_row  and cell_str(md_row[col_idx])  == "A": syms.append("MD")
        if udi_row and cell_str(udi_row[col_idx])  == "A": syms.append("UDI")
        if dom_row and cell_str(dom_row[col_idx])  == "A": syms.append("Date of Manufacture")
        if ifu_row and cell_str(ifu_row[col_idx])  == "A": syms.append("IFU")

        pn_to_symbols[pn] = syms
        log(f"  PN {pn:<12} → [{', '.join(syms) or 'none'}]  ER: {last_er}", "INFO")
        if progress_cb:
            progress_cb(f"Mapped PN: {pn}")

    log(f"  Total PNs: {len(pn_to_symbols)}", "INFO")
    return pn_to_symbols, pn_to_er


def build_pn_to_gtin(filepath: str, log, progress_cb=None):
    log(f"\n[Step 2] Reading: {os.path.basename(filepath)}", "INFO")
    if progress_cb:
        progress_cb("Loading GTIN workbook…")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "labeling" in nl and "accys" in nl:
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if "labeling" in name.lower():
                sheet_name = name
                break
    if not sheet_name:
        raise ValueError(
            f"Could not find 'Labeling Deliverables_Accys' sheet.\n"
            f"Available: {wb.sheetnames}"
        )

    log(f"  Sheet: '{sheet_name}'", "INFO")
    if progress_cb:
        progress_cb(f"Reading sheet: {sheet_name}…")

    ws = wb[sheet_name]
    pn_to_gtin: dict[str, str] = {}

    for row in ws.iter_rows(values_only=True):
        pn   = cell_str(row[2] if len(row) > 2 else None)
        gtin = cell_str(row[6] if len(row) > 6 else None)
        if gtin.upper().startswith("GTIN") and pn:
            pn_to_gtin[pn] = gtin

    log(f"  Total GTINs: {len(pn_to_gtin)}", "INFO")
    return pn_to_gtin


def parse_pdf_entries(filepath: str, log, progress_cb=None) -> list[dict]:
    try:
        import pdfplumber
    except ImportError:
        log("  [WARN] pdfplumber not installed. Run: pip install pdfplumber", "WARN")
        return []

    full_text_parts = []
    log(f"  Parsing PDF pages…", "INFO")

    with pdfplumber.open(filepath) as pdf:
        total = len(pdf.pages)
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            full_text_parts.append(text)
            if progress_cb:
                progress_cb(f"PDF page {i}/{total}…")

    full_text    = "\n".join(full_text_parts)
    id_positions = [(m.start(), m.group(1)) for m in PRD_ID_RE.finditer(full_text)]
    log(f"  Found {len(id_positions)} PRD IDs in PDF", "INFO")

    entries = []
    for i, (pos, prd_id) in enumerate(id_positions):
        end_pos = id_positions[i + 1][0] if i + 1 < len(id_positions) else len(full_text)
        payload = full_text[pos + len(prd_id):end_pos].replace("\n", " ").strip()
        payload = re.sub(r"\s{2,}", " ", payload)

        type_match = re.search(r"\b(Requirement|Heading|Information)\b", payload, re.IGNORECASE)
        entry_type = type_match.group(1) if type_match else ""

        if type_match:
            before = payload[:type_match.start()].strip()
            rat_split = re.split(r"(?=EN\s+ISO|ISO\s+\d|Clause\s+\d)", before, maxsplit=1, flags=re.IGNORECASE)
            req_text = rat_split[0].strip()
        else:
            req_text = payload

        entries.append({
            "id":   prd_id,
            "text": req_text,
            "full": payload,
            "type": entry_type,
        })

    return entries


def find_prd_per_symbol_pdf(entries: list[dict], required_symbols: list[str], er_number: str, log) -> dict[str, str]:
    per_sym: dict[str, list[str]] = {sym: [] for sym in required_symbols}

    for entry in entries:
        if entry["type"].lower() in ("heading", "information"):
            continue
        search_text = entry.get("full") or entry["text"]
        if re.search(r"not\s+applicable", search_text, re.IGNORECASE):
            continue

        sentence = f"[{entry['id']}] {entry['text'].strip()}"

        for sym in required_symbols:
            patterns = SYMBOL_PATTERNS.get(sym, [])
            if not patterns:
                continue
            if matches_any(search_text, patterns):
                if sentence not in per_sym[sym]:
                    per_sym[sym].append(sentence)
                    preview = sentence[:90] + ("…" if len(sentence) > 90 else "")
                    log(f"  {er_number}: [{sym}] {preview}", "INFO")

    return {sym: " | ".join(sentences) for sym, sentences in per_sym.items()}


def find_prd_per_symbol_excel(wb, required_symbols: list[str], er_number: str, log) -> dict[str, str]:
    per_sym: dict[str, list[str]] = {sym: [] for sym in required_symbols}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join(cell_str(v) for v in row)
            if not row_text.strip() or len(row_text) < 10:
                continue

            id_match   = PRD_ID_RE.search(row_text)
            prd_id     = id_match.group(1) if id_match else ""
            clean      = re.sub(r"\s+", " ", row_text).strip()
            entry_text = f"[{prd_id}] {clean}" if prd_id else clean

            for sym in required_symbols:
                patterns = SYMBOL_PATTERNS.get(sym, [])
                if not matches_any(row_text, patterns):
                    continue
                if entry_text not in per_sym[sym]:
                    per_sym[sym].append(entry_text)
                    preview = entry_text[:80] + ("…" if len(entry_text) > 80 else "")
                    log(f"  {er_number}/{sheet_name}: [{sym}] {preview}", "INFO")

    return {sym: " | ".join(sentences) for sym, sentences in per_sym.items()}


def merge_and_search(pn_to_symbols, pn_to_er, pn_to_gtin, er_file_map: dict[str, str], log, progress_cb=None):
    log("\n[Step 3] Merging + searching ER documents…", "INFO")
    log(f"  Uploaded ER keys: {list(er_file_map.keys())}", "INFO")

    pdf_cache: dict[str, list[dict]] = {}
    xl_cache:  dict[str, object]     = {}

    for er_num, path in er_file_map.items():
        if progress_cb:
            progress_cb(f"Pre-loading {er_num}…")
        if path.lower().endswith(".pdf"):
            log(f"  Parsing PDF: {er_num} ← {os.path.basename(path)}", "INFO")
            entries = parse_pdf_entries(path, log, progress_cb)
            pdf_cache[er_num] = entries
            log(f"  {er_num}: {len(entries)} PRD entries loaded", "INFO")
        else:
            log(f"  Loading Excel: {er_num} ← {os.path.basename(path)}", "INFO")
            xl_cache[er_num] = openpyxl.load_workbook(path, data_only=True)

    results = []
    total_pns = len(pn_to_gtin)
    for idx, (pn, gtin) in enumerate(pn_to_gtin.items(), 1):
        if progress_cb:
            progress_cb(f"Searching PN {idx}/{total_pns}: {pn}")

        symbols = pn_to_symbols.get(pn, [])
        er_num  = pn_to_er.get(pn, "")

        if not symbols:
            log(f"  [WARN] PN {pn}: in FRM5816 but not found in ER 2253454", "WARN")

        prd_by_sym: dict[str, str] = {sym: "" for sym in ALL_SYMBOLS}

        if er_num and symbols:
            if er_num in pdf_cache:
                log(f"  PN {pn}: searching PDF {er_num} for {symbols}", "INFO")
                found = find_prd_per_symbol_pdf(pdf_cache[er_num], symbols, er_num, log)
                prd_by_sym.update(found)
            elif er_num in xl_cache:
                log(f"  PN {pn}: searching Excel {er_num} for {symbols}", "INFO")
                found = find_prd_per_symbol_excel(xl_cache[er_num], symbols, er_num, log)
                prd_by_sym.update(found)
            else:
                log(f"  [WARN] PN {pn}: ER number '{er_num}' not in uploaded docs {list(er_file_map.keys())}", "WARN")

        row = {
            "GTIN":                gtin,
            "PN":                  pn,
            "MD":                  "YES" if "MD"                   in symbols else "",
            "UDI":                 "YES" if "UDI"                  in symbols else "",
            "Date of Manufacture": "YES" if "Date of Manufacture"  in symbols else "",
            "IFU":                 "YES" if "IFU"                  in symbols else "",
            "Symbols Required":    ", ".join(symbols),
            "ER Number":           er_num,
        }
        for sym in ALL_SYMBOLS:
            row[prd_col(sym)] = prd_by_sym.get(sym, "")

        results.append(row)

    results.sort(key=lambda r: r["GTIN"])
    return results


BASE_FIELDS = ["GTIN", "PN", "MD", "UDI", "Date of Manufacture", "IFU",
               "Symbols Required", "ER Number"]
PRD_FIELDS  = [prd_col(sym) for sym in ALL_SYMBOLS]
ALL_FIELDS  = BASE_FIELDS + PRD_FIELDS

BASE_WIDTHS = [22, 14, 6, 6, 22, 6, 30, 14]
PRD_WIDTHS  = [60, 60, 60, 60]
ALL_WIDTHS  = BASE_WIDTHS + PRD_WIDTHS


def save_excel(results, output_path: str, log):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PRD Mapping"

    header_fill  = PatternFill("solid", fgColor="1E3A8A")
    header_font  = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align   = Alignment(wrap_text=True, vertical="top")
    prd_fill     = PatternFill("solid", fgColor="EFF6FF")

    for ci, field in enumerate(ALL_FIELDS, start=1):
        cell = ws.cell(row=1, column=ci, value=field)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    for ri, row in enumerate(results, start=2):
        for ci, field in enumerate(ALL_FIELDS, start=1):
            val  = row.get(field, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            if field in PRD_FIELDS:
                cell.alignment = wrap_align
                if val:
                    cell.fill = prd_fill
            elif field in ("MD", "UDI", "Date of Manufacture", "IFU") and val == "YES":
                cell.font = Font(bold=True, color="166534")

    for ci, width in enumerate(ALL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
    log(f"\n[Saved] {output_path}", "INFO")


def save_csv(results, output_path: str, log):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ALL_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    log(f"[Saved] {output_path}", "INFO")


# ══════════════════════════════════════════════════════════════════════════════
# GUI — Mapper Window
# ══════════════════════════════════════════════════════════════════════════════

class MapperWindow(tk.Toplevel):

    PROCESS_STEPS = [
        "Load Impact Assessment",
        "Load Design Sheet (FRM5816)",
        "Load ER Documents",
        "Extract PRD entries per symbol",
        "Build results table",
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.title("Impact Assessment & PRD Mapper")
        self.configure(bg=COLORS["bg"])
        self.resizable(True, True)

        self.file1_path  = tk.StringVar()
        self.file2_path  = tk.StringVar()
        self.er_files: dict[str, str] = {}
        self.results     = []
        self._popup      = None

        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = 1300, 850
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # ── UI Construction ────────────────────────────────────────────────────
    def _build_ui(self):
        self._setup_styles()

        # Header
        hdr = tk.Frame(self, bg=COLORS["header_bg"], padx=16, pady=12)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Impact Assessment & PRD Mapper", bg=COLORS["header_bg"],
                 fg="white", font=FONT_TITLE).pack(side="left")
        tk.Label(hdr, text="  ·  Step 1 of 3", bg=COLORS["header_bg"],
                 fg="#93C5FD", font=FONT_BODY).pack(side="left")

        # Main paned
        paned = tk.PanedWindow(self, orient="vertical", bg=COLORS["bg"],
                               sashwidth=6, sashrelief="flat", sashpad=2)
        paned.pack(fill="both", expand=True, padx=10, pady=10)

        top = tk.Frame(paned, bg=COLORS["bg"])
        bot = tk.Frame(paned, bg=COLORS["bg"])
        paned.add(top, minsize=280)
        paned.add(bot, minsize=240)

        self._build_top_panel(top)
        self._build_bottom_panel(bot)

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        style.configure("Dark.TLabelframe",
                         background=COLORS["bg_card"],
                         relief="flat",
                         borderwidth=1,
                         bordercolor=COLORS["border"])
        style.configure("Dark.TLabelframe.Label",
                         background=COLORS["bg_card"],
                         foreground=COLORS["accent"],
                         font=FONT_HEADING)
        style.configure("Dark.TNotebook",
                         background=COLORS["bg"],
                         tabmargins=[2, 5, 2, 0])
        style.configure("Dark.TNotebook.Tab",
                         background=COLORS["bg_card"],
                         foreground=COLORS["text_muted"],
                         padding=[12, 6],
                         font=FONT_BODY)
        style.map("Dark.TNotebook.Tab",
                  background=[("selected", COLORS["accent"])],
                  foreground=[("selected", "white")])
        style.configure("Dark.Horizontal.TProgressbar",
                         troughcolor=COLORS["border"],
                         background=COLORS["accent"],
                         thickness=6)
        style.configure("Dark.Treeview",
                         background=COLORS["bg_card"],
                         foreground=COLORS["text"],
                         fieldbackground=COLORS["bg_card"],
                         rowheight=26,
                         font=FONT_MONO)
        style.configure("Dark.Treeview.Heading",
                         background=COLORS["header_bg"],
                         foreground="white",
                         font=FONT_SMALL,
                         relief="flat")
        style.map("Dark.Treeview", background=[("selected", COLORS["accent"])])

    def _section(self, parent, title):
        f = ttk.LabelFrame(parent, text=f"  {title}  ", style="Dark.TLabelframe",
                            padding=(12, 8))
        f.configure(labelanchor="n")
        return f

    def _build_top_panel(self, parent):
        # File 1
        f1 = self._section(parent, "Upload Impact Assessment")
        f1.pack(fill="x", pady=(0, 6))
        self._file_row(f1, self.file1_path,
                       "Excel — Impact Assessment (symbol gap matrix with FGN rows)",
                       self._browse_file1)

        # File 2
        f2 = self._section(parent, "Upload Design Sheet (FRM5816)")
        f2.pack(fill="x", pady=(0, 6))
        self._file_row(f2, self.file2_path,
                       "Excel — Design Sheet (FRM5816, Labeling Deliverables_Accys tab)",
                       self._browse_file2)

        # ER Documents
        er = self._section(parent, "Upload ER Documents (for PRD Extraction)")
        er.pack(fill="x", pady=(0, 6))

        er_top = tk.Frame(er, bg=COLORS["bg_card"])
        er_top.pack(fill="x", pady=(0, 4))
        tk.Label(er_top, text='Filename must contain the ER number  e.g.  "ER2253452_PRD.pdf"',
                 bg=COLORS["bg_card"], fg=COLORS["text_muted"], font=FONT_SMALL).pack(side="left")

        btn_er_add = tk.Button(er_top, text="+ Add ER Files", command=self._browse_er_files,
                               bg=COLORS["accent3"], fg="white", activebackground=COLORS["accent3"],
                               activeforeground="white", relief="flat", cursor="hand2",
                               font=FONT_SMALL, padx=10, pady=4, bd=0)
        btn_er_add.pack(side="right")

        er_list_frame = tk.Frame(er, bg=COLORS["bg_card"])
        er_list_frame.pack(fill="x")

        self.er_listbox = tk.Listbox(er_list_frame, height=4,
                                      bg=COLORS["bg_input"], fg=COLORS["text"],
                                      selectbackground=COLORS["accent"],
                                      font=FONT_MONO, relief="flat",
                                      highlightthickness=1,
                                      highlightbackground=COLORS["border"])
        vsb = ttk.Scrollbar(er_list_frame, orient="vertical", command=self.er_listbox.yview)
        self.er_listbox.configure(yscrollcommand=vsb.set)
        self.er_listbox.pack(side="left", fill="x", expand=True)
        vsb.pack(side="right", fill="y")

        tk.Button(er, text="Remove Selected", command=self._remove_er_file,
                  bg=COLORS["danger"], fg="white", activebackground=COLORS["danger"],
                  activeforeground="white", relief="flat", cursor="hand2",
                  font=FONT_SMALL, padx=8, pady=3, bd=0).pack(anchor="e", pady=(4, 0))

        # Action buttons + status
        btn_row = tk.Frame(parent, bg=COLORS["bg"], pady=6)
        btn_row.pack(fill="x")

        run_btn = tk.Button(btn_row, text="▶  Run PRD Mapping",
                            command=self._run,
                            bg=COLORS["accent"], fg="white",
                            activebackground=_lighten(COLORS["accent"]),
                            activeforeground="white", relief="flat",
                            cursor="hand2", font=FONT_HEADING,
                            padx=18, pady=10, bd=0)
        run_btn.pack(side="left", padx=(0, 10))

        self.save_btn = tk.Button(btn_row, text="⬇  Export Results (.xlsx / .csv)",
                                   command=self._save,
                                   bg=COLORS["accent2"], fg="white",
                                   activebackground=_lighten(COLORS["accent2"]),
                                   activeforeground="white", relief="flat",
                                   cursor="hand2", font=FONT_BODY,
                                   padx=14, pady=10, bd=0,
                                   state="disabled",
                                   disabledforeground="#4B5563")
        self.save_btn.pack(side="left")

        self.status_var = tk.StringVar(value="Ready — upload files and click Run PRD Mapping.")
        tk.Label(btn_row, textvariable=self.status_var,
                 bg=COLORS["bg"], fg=COLORS["text_muted"], font=FONT_SMALL).pack(side="left", padx=14)

        # Mini progress bar (visible during processing)
        self._mini_pb = ttk.Progressbar(parent, style="Dark.Horizontal.TProgressbar",
                                         orient="horizontal", mode="indeterminate")
        self._mini_pb.pack(fill="x", pady=(0, 2))

    def _file_row(self, parent, var, hint, cmd):
        row = tk.Frame(parent, bg=COLORS["bg_card"])
        row.pack(fill="x", pady=(0, 2))

        entry = tk.Entry(row, textvariable=var, state="readonly",
                         bg=COLORS["bg_input"], fg=COLORS["text"],
                         readonlybackground=COLORS["bg_input"],
                         relief="flat", font=FONT_MONO,
                         highlightthickness=1, highlightbackground=COLORS["border"])
        entry.pack(side="left", fill="x", expand=True, padx=(0, 6))

        tk.Button(row, text="Browse…", command=cmd,
                  bg=COLORS["bg_card"], fg=COLORS["text"],
                  activebackground=COLORS["border"], activeforeground=COLORS["text"],
                  relief="flat", cursor="hand2", font=FONT_SMALL,
                  padx=10, pady=5, bd=0,
                  highlightthickness=1, highlightbackground=COLORS["border"]).pack(side="left")

        tk.Label(parent, text=hint, bg=COLORS["bg_card"],
                 fg=COLORS["text_muted"], font=FONT_SMALL).pack(anchor="w")

    def _build_bottom_panel(self, parent):
        nb = ttk.Notebook(parent, style="Dark.TNotebook")
        nb.pack(fill="both", expand=True)
        self._nb = nb

        # ── Log tab ──
        log_frame = tk.Frame(nb, bg=COLORS["log_bg"])
        nb.add(log_frame, text="  Processing Log  ")

        self.log_text = scrolledtext.ScrolledText(
            log_frame, bg=COLORS["log_bg"], fg=COLORS["log_text"],
            insertbackground=COLORS["text"],
            font=FONT_MONO, state="disabled", relief="flat", wrap="word",
        )
        self.log_text.pack(fill="both", expand=True, padx=1, pady=1)
        self.log_text.tag_config("INFO",  foreground=COLORS["log_text"])
        self.log_text.tag_config("WARN",  foreground=COLORS["warn"])
        self.log_text.tag_config("ERROR", foreground=COLORS["danger"])

        # ── Results tab ──
        res_frame = tk.Frame(nb, bg=COLORS["bg"])
        nb.add(res_frame, text="  Results Table  ")
        self._build_results_table(res_frame)

    def _build_results_table(self, parent):
        tree_cols = ("GTIN", "PN", "MD", "UDI", "Date of Mfg", "IFU",
                     "ER Number", "Symbols Required",
                     "PRD for MD", "PRD for UDI", "PRD for DOM", "PRD for IFU")
        tree_widths = (150, 100, 45, 45, 110, 45, 100, 170, 260, 260, 260, 260)

        tf = tk.Frame(parent, bg=COLORS["bg"])
        tf.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(tf, columns=tree_cols, show="headings",
                                  selectmode="browse", style="Dark.Treeview")
        vsb = ttk.Scrollbar(tf, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tf.grid_rowconfigure(0, weight=1)
        tf.grid_columnconfigure(0, weight=1)

        for col, width in zip(tree_cols, tree_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, minwidth=40)

        self.tree.tag_configure("odd",      background="#1E293B", foreground=COLORS["text"])
        self.tree.tag_configure("even",     background="#162032", foreground=COLORS["text"])
        self.tree.tag_configure("has_prd",  background="#1C3A2E", foreground="#4ADE80")
        self.tree.tag_configure("no_prd",   background="#3A1C1C", foreground="#FCA5A5")

        # Double-click to expand row
        self.tree.bind("<Double-1>", self._on_row_double_click)

    # ── Browsing ──────────────────────────────────────────────────────────

    def _refocus(self):
        """Re-raise this window after a file dialog closes on Windows."""
        self.lift()
        self.focus_force()
        self.attributes("-topmost", True)
        self.after(150, lambda: self.attributes("-topmost", False))

    def _browse_file1(self):
        p = filedialog.askopenfilename(parent=self, 
            title="Select Impact Assessment Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            self.file1_path.set(p)
            self._log("INFO", f"File 1 set: {os.path.basename(p)}")

        self._refocus()
    def _browse_file2(self):
        p = filedialog.askopenfilename(parent=self, 
            title="Select Design Sheet (FRM5816)",
            filetypes=[("Excel files", "*.xlsx *.xls")])
        if p:
            self.file2_path.set(p)
            self._log("INFO", f"File 2 set: {os.path.basename(p)}")

        self._refocus()
    def _browse_er_files(self):
        paths = filedialog.askopenfilenames(parent=self, 
            title="Select ER Documents (PDF or Excel) — for PRD Extraction",
            filetypes=[
                ("ER documents", "*.pdf *.xlsx *.xls"),
                ("PDF files",    "*.pdf"),
                ("Excel files",  "*.xlsx *.xls"),
            ])
        for p in paths:
            name = os.path.basename(p)
            m = re.search(r"ER\s*(\d{7,})", name, re.IGNORECASE)
            if not m:
                messagebox.showwarning(
                    "ER Number Not Found",
                    f"Could not detect ER number in:\n{name}\n\n"
                    f"Filename must contain the ER number, e.g. ER2253452_PRD.pdf",
                    parent=self)
                continue
            er_num = f"ER{m.group(1)}"
            ext = "(PDF)" if p.lower().endswith(".pdf") else "(Excel)"
            if er_num not in self.er_files:
                self.er_listbox.insert("end", f"{er_num}  {ext}  ←  {name}")
            self.er_files[er_num] = p
            self._log("INFO", f"ER file added: {er_num} {ext}")
        self._refocus()
    def _remove_er_file(self):
        for idx in reversed(self.er_listbox.curselection()):
            line   = self.er_listbox.get(idx)
            er_num = line.split("  ")[0].strip()
            self.er_files.pop(er_num, None)
            self.er_listbox.delete(idx)

    # ── Logging ───────────────────────────────────────────────────────────
    def _log(self, level: str, msg: str):
        self.log_text.configure(state="normal")
        prefix = {"INFO": "› ", "WARN": "⚠ ", "ERROR": "✗ "}.get(level, "  ")
        self.log_text.insert("end", prefix + msg + "\n", level)
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update_idletasks()

    def _clear_log(self):
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")

    # ── Run / Process ─────────────────────────────────────────────────────
    def _run(self):
        if not self.file1_path.get():
            messagebox.showerror("Missing File", "Please select the Impact Assessment Excel file.", parent=self)
            return
        if not self.file2_path.get():
            messagebox.showerror("Missing File", "Please select the Design Sheet (FRM5816).", parent=self)
            return

        self._clear_log()
        self.save_btn.configure(state="disabled")
        self.results = []
        self.status_var.set("Processing…")
        self._nb.select(0)
        self._mini_pb.start(15)

        # Show progress popup
        self._popup = ProgressPopup(self, title="PRD Mapping", steps=self.PROCESS_STEPS)
        self._popup.set_step(0)

        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        try:
            def log(msg, level="INFO"):
                self.after(0, self._log, level, msg)

            def progress_cb(detail):
                if self._popup and self._popup.winfo_exists():
                    self.after(0, self._popup.set_detail, detail)

            # Step 1 — Symbol matrix
            self.after(0, self._popup.set_step, 0, "Reading symbol matrix…")
            pn_to_symbols, pn_to_er = build_pn_to_symbols(
                self.file1_path.get(), log, progress_cb)

            # Step 2 — GTIN mapping
            self.after(0, self._popup.set_step, 1, "Reading GTIN form…")
            pn_to_gtin = build_pn_to_gtin(self.file2_path.get(), log, progress_cb)

            # Step 3 — Pre-load ER docs
            self.after(0, self._popup.set_step, 2, "Loading ER documents…")

            # Step 4 — Search (merge_and_search handles steps 3+4 internally)
            self.after(0, self._popup.set_step, 3, "Searching PRD entries…")
            results = merge_and_search(
                pn_to_symbols, pn_to_er, pn_to_gtin, self.er_files, log, progress_cb)

            # Step 5 — Build table
            self.after(0, self._popup.set_step, 4, "Building results table…")
            self.results = results
            self.after(0, self._on_done, results)

        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            self.after(0, self._log, "ERROR", str(e))
            self.after(0, self._log, "ERROR", tb)
            self.after(0, self.status_var.set, "Error — see log.")
            self.after(0, self._mini_pb.stop)
            if self._popup:
                self.after(0, self._popup.error, str(e)[:80])

    def _on_done(self, results):
        self._mini_pb.stop()
        self.status_var.set(f"Done — {len(results)} rows.")
        self.save_btn.configure(state="normal")
        self._populate_table(results)
        self._nb.select(1)
        with_prd = sum(1 for r in results if any(r.get(prd_col(s)) for s in ALL_SYMBOLS))
        self._log("INFO", f"\nComplete — {len(results)} GTINs, {with_prd} with PRD entries.")
        if self._popup:
            self._popup.finish(f"Complete — {len(results)} GTINs processed.")

        # Build verification summary and ask where to save it
        summary = self._build_verify_summary(results)
        if any(entry["gtin_count"] > 0 for entry in summary):
            save_path = self._save_verification_list(summary)
            self._show_verify_popup(summary, save_path)

        # ── Automation: notify Label Verifier of required symbols ────────────
        if APP_STATE is not None:
            APP_STATE["required_symbols"]   = [s["symbol"] for s in summary if s.get("gtin_count", 0) > 0]
            APP_STATE["verify_list"]        = summary
            # Total unique GTINs = deduplicated union across all per-symbol GTIN lists
            _all_unique = {g for s in summary for g in s.get("gtins", [])}
            APP_STATE["total_unique_gtins"] = len(_all_unique)
        fire_event("symbols_ready", summary)

    def _populate_table(self, results):
        for item in self.tree.get_children():
            self.tree.delete(item)

        def trunc(text, n=120):
            return text[:n] + ("…" if len(text) > n else "")

        for i, r in enumerate(results):
            has_prd = any(r.get(prd_col(s)) for s in ALL_SYMBOLS)
            tag = "has_prd" if has_prd else ("odd" if i % 2 else "even")
            self.tree.insert("", "end", tag=tag, values=(
                r["GTIN"], r["PN"],
                r["MD"], r["UDI"], r["Date of Manufacture"], r["IFU"],
                r["ER Number"], r["Symbols Required"],
                trunc(r.get(prd_col("MD"), "")),
                trunc(r.get(prd_col("UDI"), "")),
                trunc(r.get(prd_col("Date of Manufacture"), "")),
                trunc(r.get(prd_col("IFU"), "")),
            ))

    def _on_row_double_click(self, event):
        """Show full PRD content in a detail popup."""
        sel = self.tree.selection()
        if not sel:
            return
        idx = self.tree.index(sel[0])
        if idx >= len(self.results):
            return
        r = self.results[idx]
        self._show_detail_window(r)

    def _show_detail_window(self, r):
        win = tk.Toplevel(self)
        win.title(f"Detail — {r['PN']} / {r['GTIN']}")
        win.configure(bg=COLORS["bg"])
        win.geometry("700x500")

        hdr = tk.Frame(win, bg=COLORS["header_bg"], pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text=f"PN: {r['PN']}   GTIN: {r['GTIN']}   ER: {r['ER Number']}",
                 bg=COLORS["header_bg"], fg="white", font=FONT_HEADING).pack(padx=14)

        body = scrolledtext.ScrolledText(
            win, bg=COLORS["log_bg"], fg=COLORS["text"],
            font=FONT_MONO, relief="flat", wrap="word")
        body.pack(fill="both", expand=True, padx=8, pady=8)

        body.tag_config("head", foreground=COLORS["accent"], font=("Segoe UI", 9, "bold"))
        body.tag_config("val",  foreground=COLORS["text"])
        body.tag_config("yes",  foreground=COLORS["accent2"], font=("Segoe UI", 9, "bold"))

        fields = [
            ("GTIN",                r["GTIN"]),
            ("PN",                  r["PN"]),
            ("ER Number",           r["ER Number"]),
            ("Symbols Required",    r["Symbols Required"]),
            ("MD",                  r["MD"]),
            ("UDI",                 r["UDI"]),
            ("Date of Manufacture", r["Date of Manufacture"]),
            ("IFU",                 r["IFU"]),
        ]
        for label, val in fields:
            body.insert("end", f"{label}:  ", "head")
            tag = "yes" if val == "YES" else "val"
            body.insert("end", f"{val}\n", tag)

        body.insert("end", "\n── PRD References ──────────────────────────────\n\n", "head")
        for sym in ALL_SYMBOLS:
            col = prd_col(sym)
            body.insert("end", f"{col}:\n", "head")
            body.insert("end", r.get(col, "(none)") + "\n\n", "val")

        body.configure(state="disabled")

    # ── Verification list ──────────────────────────────────────────────────

    def _build_verify_summary(self, results: list[dict]) -> list[dict]:
        """
        For each symbol in ALL_SYMBOLS, collect every GTIN that requires it.
        Returns a list of dicts: {symbol, gtin_count, gtins: [...]}.
        """
        summary = []
        for sym in ALL_SYMBOLS:
            gtins = [
                r["GTIN"] for r in results
                if r.get(sym, "").strip().upper() == "YES"
            ]
            summary.append({
                "symbol":     sym,
                "gtin_count": len(gtins),
                "gtins":      gtins,
            })
        return summary

    def _save_verification_list(self, summary: list[dict]) -> str | None:
        """Ask the user where to save symbols_to_verify.json and write it."""
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Symbols-to-Verify List",
            defaultextension=".json",
            initialfile="symbols_to_verify.json",
            filetypes=[("JSON file", "*.json"), ("All files", "*.*")],
        )
        if not path:
            return None

        payload = {
            "generated_by":      "GTIN Symbol Mapper",
            "generated_at":      datetime.now().isoformat(timespec="seconds"),
            "total_gtins":       len(self.results),
            "symbols_to_verify": [
                s for s in summary if s["gtin_count"] > 0
            ],
        }
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, indent=2, ensure_ascii=False)

        self._log("INFO", f"[Saved] Verification list → {path}")
        return path

    def _show_verify_popup(self, summary: list[dict], save_path: str | None):
        """Dark-themed popup listing which symbols need label verification."""
        needed = [s for s in summary if s["gtin_count"] > 0]
        if not needed:
            return

        pop = tk.Toplevel(self)
        pop.title("Symbols to Verify")
        pop.configure(bg=COLORS["bg"])
        pop.resizable(False, False)
        pop.grab_set()

        # ── Header ───────────────────────────────────────────────────────
        # Total unique GTINs across all required symbols
        _all_unique_gtins = {g for s in needed for g in s.get("gtins", [])}
        _total_unique = len(_all_unique_gtins)

        hdr = tk.Frame(pop, bg=COLORS["header_bg"], padx=20, pady=12)
        hdr.pack(fill="x")

        hdr_top = tk.Frame(hdr, bg=COLORS["header_bg"])
        hdr_top.pack(fill="x")
        tk.Label(hdr_top, text="🔎  Symbols to Verify in EVO MDD Verification Tool",
                 bg=COLORS["header_bg"], fg="white",
                 font=FONT_HEADING).pack(side="left")

        # Prominent GTIN total badge on the right
        badge_frame = tk.Frame(hdr_top, bg="#1E5FAD", padx=12, pady=4)
        badge_frame.pack(side="right")
        tk.Label(badge_frame,
                 text=str(_total_unique),
                 bg="#1E5FAD", fg="white",
                 font=("Segoe UI", 16, "bold")).pack(side="left")
        tk.Label(badge_frame,
                 text="  GTINs to Verify",
                 bg="#1E5FAD", fg="#BFDBFE",
                 font=FONT_BODY).pack(side="left")

        tk.Label(hdr,
                 text="The following symbols were identified in your Impact Assessment as requiring label verification.",
                 bg=COLORS["header_bg"], fg="#93C5FD",
                 font=FONT_SMALL).pack(anchor="w", pady=(4, 0))

        # ── Symbol cards ─────────────────────────────────────────────────
        body = tk.Frame(pop, bg=COLORS["bg"], padx=18, pady=14)
        body.pack(fill="x")

        for entry in needed:
            card = tk.Frame(body, bg=COLORS["bg_card"],
                            highlightthickness=1,
                            highlightbackground=COLORS["accent"])
            card.pack(fill="x", pady=4)

            accent_bar = tk.Frame(card, bg=COLORS["accent"], width=4)
            accent_bar.pack(side="left", fill="y")

            inner = tk.Frame(card, bg=COLORS["bg_card"], padx=12, pady=8)
            inner.pack(side="left", fill="both", expand=True)
            inner.columnconfigure(1, weight=1)

            tk.Label(inner, text=entry["symbol"],
                     bg=COLORS["bg_card"], fg=COLORS["text"],
                     font=FONT_HEADING, anchor="w").grid(
                row=0, column=0, sticky="w")

            count_lbl = tk.Label(
                inner,
                text=f"{entry['gtin_count']} GTIN{'s' if entry['gtin_count'] != 1 else ''} require this",
                bg=COLORS["bg_card"], fg=COLORS["accent2"],
                font=FONT_SMALL,
            )
            count_lbl.grid(row=0, column=1, sticky="e")

            gtin_preview = ", ".join(entry["gtins"][:4])
            if len(entry["gtins"]) > 4:
                gtin_preview += f"  …+{len(entry['gtins'])-4} more"
            tk.Label(inner, text=gtin_preview,
                     bg=COLORS["bg_card"], fg=COLORS["text_muted"],
                     font=FONT_MONO, anchor="w").grid(
                row=1, column=0, columnspan=2, sticky="w", pady=(2, 0))

        # ── Save path notice ─────────────────────────────────────────────
        notice_frame = tk.Frame(pop, bg=COLORS["bg_card"], padx=18, pady=8)
        notice_frame.pack(fill="x")
        if save_path:
            notice_text = f"✔  Saved to: {save_path}"
            notice_fg   = COLORS["success"]
        else:
            notice_text = "⚠  File was not saved (dialog cancelled)."
            notice_fg   = COLORS["warn"]
        tk.Label(notice_frame, text=notice_text,
                 bg=COLORS["bg_card"], fg=notice_fg,
                 font=FONT_SMALL, wraplength=520, justify="left").pack(
            anchor="w")

        # ── Instruction ───────────────────────────────────────────────────
        instr = tk.Frame(pop, bg=COLORS["bg"], padx=18, pady=6)
        instr.pack(fill="x")
        tk.Label(instr,
                 text="Open the EVO MDD Verification Tool (Step 3) to verify these symbols against your device label.",
                 bg=COLORS["bg"], fg=COLORS["text_muted"],
                 font=FONT_SMALL, wraplength=520, justify="left").pack(anchor="w")

        # ── Close button ──────────────────────────────────────────────────
        btn_row = tk.Frame(pop, bg=COLORS["bg"], pady=12)
        btn_row.pack()
        tk.Button(btn_row, text="OK  —  Got it",
                  command=pop.destroy,
                  bg=COLORS["accent"], fg="white",
                  activebackground=_lighten(COLORS["accent"]),
                  activeforeground="white",
                  relief="flat", cursor="hand2",
                  font=FONT_BODY, padx=24, pady=8, bd=0).pack()

        pop.update_idletasks()
        w = pop.winfo_reqwidth()
        h = pop.winfo_reqheight()
        sw, sh = pop.winfo_screenwidth(), pop.winfo_screenheight()
        pop.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # ── Save ──────────────────────────────────────────────────────────────
    def _save(self):
        if not self.results:
            messagebox.showinfo("No Results", "Run processing first.", parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self,
            title="Save Output",
            defaultextension=".xlsx",
            initialfile="gtin_symbol_output_with_prd.xlsx",
            filetypes=[("Excel file", "*.xlsx"), ("CSV file", "*.csv")])
        if not path:
            return
        try:
            log_fn = lambda m, lvl="INFO": self._log(lvl, m)
            if path.endswith(".csv"):
                save_csv(self.results, path, log_fn)
            else:
                save_excel(self.results, path, log_fn)
            messagebox.showinfo("Saved", f"Output saved:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Save Error", str(e), parent=self)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    win = MapperWindow(root)
    win.mainloop()